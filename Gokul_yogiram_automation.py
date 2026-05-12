import streamlit as st
# Version: 2.1.0-standalone
import pandas as pd
import numpy as np
import re
from io import BytesIO
from datetime import datetime, date, timedelta
import os
import sys
import time
import sys
import os

# Add pharma_forecast_app to sys.path to allow local imports within that folder
current_dir = os.path.dirname(os.path.abspath(__file__))
forecast_app_path = os.path.join(current_dir, "pharma_forecast_app")
if os.path.exists(forecast_app_path) and forecast_app_path not in sys.path:
    sys.path.append(forecast_app_path)

try:
    from modules.forecast_module import render_forecast_module
except ImportError:
    render_forecast_module = None

# =========================================================
# PHARMA FORECAST STANDALONE ENGINE (MODULAR REDIRECT)
# =========================================================


# ------------------ UTILS ------------------
def load_data_with_progress(uploaded_file):
    with st.spinner("🚀 Analyzing Data Structure..."):
        time.sleep(0.5) # Aesthetic pause
        if uploaded_file.name.endswith(".csv"):
            return pd.read_csv(uploaded_file)
        else:
            return pd.read_excel(uploaded_file)

def save_df_to_excel_with_format(df, sheet_name="Data", index=True):
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=index, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Colors (Orange Accent 6 Lighter 40% -> #F4B084)
        header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format header
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format data and add borders
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for i, col in enumerate(worksheet.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            worksheet.column_dimensions[column].width = max_length + 2

    return out.getvalue()

def save_multi_df_to_excel_with_format(df_dict):
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            
            header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                           min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
            
            for i, col in enumerate(worksheet.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2

    return out.getvalue()

# ------------------ PAGE CONFIG ------------------
st.set_page_config(
    page_title="Yogiram Automation - Gokul",
    page_icon="⚙️",
    layout="wide",  # force wide layout
    initial_sidebar_state="collapsed"
)

# ------------------ GLOBAL PREMIUM STYLING ------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700;800&family=Plus+Jakarta+Sans:wght@300;400;600;700&display=swap');

/* ============ ROOT VARIABLES ============ */
:root {
    --primary: #4f46e5;
    --primary-glow: rgba(79, 70, 229, 0.4);
    --secondary: #06b6d4;
    --accent: #f43f5e;
    --gold: #f59e0b;
    --bg: #030305;
    --surface: rgba(255,255,255,0.03);
    --border: rgba(255,255,255,0.08);
    --text: #f8fafc;
    --muted: #94a3b8;
}

/* ============ APP BACKGROUND ============ */
.stApp {
    background: var(--bg) !important;
    background-image: 
        radial-gradient(circle at 15% 50%, rgba(79, 70, 229, 0.08), transparent 25%),
        radial-gradient(circle at 85% 30%, rgba(6, 182, 212, 0.08), transparent 25%),
        radial-gradient(circle at 50% 100%, rgba(244, 63, 94, 0.06), transparent 30%) !important;
    background-attachment: fixed !important;
}

[data-testid="stAppViewContainer"] { background: transparent; }
[data-testid="stHeader"] { background: transparent; }

/* ============ TYPOGRAPHY ============ */
.main p, .main span, .main div, .main label {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    color: var(--text); /* Remove !important to let widgets use their internal colors */
}
h1, h2, h3, h4 {
    font-family: 'Outfit', sans-serif !important;
    color: #fff !important;
}

/* ============ HIDE CHROME ============ */

footer    { visibility: hidden; }


/* ============ SCROLLBAR ============ */
::-webkit-scrollbar { width: 5px; }
::-webkit-scrollbar-track { background: #0a0c10; }
::-webkit-scrollbar-thumb { background: var(--primary); border-radius: 99px; }

/* ============ LOGIN PAGE ============ */
/* Target the middle column on the login page */
[data-testid="stVerticalBlock"]:has(> [data-testid="stImage"]) {
    background: rgba(255,255,255,0.04) !important;
    border: 1px solid rgba(0,242,255,0.2) !important;
    border-radius: 28px !important;
    padding: 2.5rem !important;
    box-shadow: 0 0 60px rgba(0,242,255,0.06), 0 25px 50px rgba(0,0,0,0.5) !important;
    backdrop-filter: blur(20px) !important;
}

/* ============ INPUT FIELDS ============ */
.stTextInput > div > div > input {
    background: rgba(255,255,255,0.04) !important;
    border: 1px solid var(--border) !important;
    border-radius: 14px !important;
    color: #fff !important;
    font-size: 1rem !important;
    transition: border-color 0.25s, box-shadow 0.25s !important;
}
.stTextInput > div > div > input:focus {
    border-color: var(--primary) !important;
    box-shadow: 0 0 0 3px rgba(0,242,255,0.15) !important;
    outline: none !important;
}
.stTextInput label {
    color: var(--muted) !important;
    font-size: 0.85rem !important;
    letter-spacing: 0.05em !important;
}

/* ============ BUTTONS — LOGIN ============ */
[data-testid="stVerticalBlock"]:has(> [data-testid="stImage"]) .stButton > button {
    width: 100% !important;
    background: linear-gradient(135deg, var(--primary), rgba(0,242,255,0.5)) !important;
    color: #000 !important;
    font-weight: 800 !important;
    font-size: 1rem !important;
    border: none !important;
    border-radius: 14px !important;
    padding: 0.75rem !important;
    letter-spacing: 0.1em !important;
    transition: all 0.3s !important;
    box-shadow: 0 0 20px rgba(0,242,255,0.35) !important;
    min-height: auto !important;
}
[data-testid="stVerticalBlock"]:has(> [data-testid="stImage"]) .stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 0 35px rgba(0,242,255,0.6) !important;
}

/* ============ DASHBOARD GRID BUTTONS ============ */
.main .stButton > button {
    width: 100% !important;
    height: 85px !important;
    min-height: 85px !important;
    max-height: 85px !important;
    display: flex !important;
    align-items: center !important;
    justify-content: flex-start !important; /* Left alignment for modern card look */
    background: linear-gradient(180deg, rgba(255,255,255,0.06) 0%, rgba(255,255,255,0.01) 100%) !important;
    border: 1px solid rgba(255, 255, 255, 0.08) !important;
    border-top: 1px solid rgba(255, 255, 255, 0.15) !important; /* Top highlight */
    border-radius: 16px !important;
    color: #f1f5f9 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1.05rem !important;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06), inset 0 1px 0 rgba(255,255,255,0.1) !important;
    backdrop-filter: blur(12px) !important;
    -webkit-backdrop-filter: blur(12px) !important;
    padding: 0 1.5rem !important;
    white-space: nowrap !important;
    text-shadow: none !important;
}
.main .stButton > button > div > p {
    display: flex !important;
    align-items: center !important;
    gap: 12px !important;
    margin: 0 !important;
}
.main .stButton > button:hover {
    background: linear-gradient(180deg, rgba(255,255,255,0.08) 0%, rgba(255,255,255,0.02) 100%) !important;
    transform: translateY(-3px) scale(1.01) !important;
    border-color: rgba(255, 255, 255, 0.2) !important;
    border-top: 1px solid rgba(255, 255, 255, 0.3) !important;
    box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.2), 0 10px 10px -5px rgba(0, 0, 0, 0.1), inset 0 1px 0 rgba(255,255,255,0.2) !important;
    color: #ffffff !important;
}
/* Premium subtle glows per column instead of harsh neon */
div[data-testid="column"]:nth-child(1) .main .stButton > button:hover {
    box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.2), 0 0 20px rgba(79, 70, 229, 0.4), inset 0 1px 0 rgba(255,255,255,0.2) !important;
    border-color: rgba(79, 70, 229, 0.5) !important;
}
div[data-testid="column"]:nth-child(2) .main .stButton > button:hover {
    box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.2), 0 0 20px rgba(6, 182, 212, 0.3), inset 0 1px 0 rgba(255,255,255,0.2) !important;
    border-color: rgba(6, 182, 212, 0.5) !important;
}
div[data-testid="column"]:nth-child(3) .main .stButton > button:hover {
    box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.2), 0 0 20px rgba(244, 63, 94, 0.3), inset 0 1px 0 rgba(255,255,255,0.2) !important;
    border-color: rgba(244, 63, 94, 0.5) !important;
}
div[data-testid="column"]:nth-child(4) .main .stButton > button:hover {
    box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.2), 0 0 20px rgba(245, 158, 11, 0.3), inset 0 1px 0 rgba(255,255,255,0.2) !important;
    border-color: rgba(245, 158, 11, 0.5) !important;
}

/* ============ SECTION LABELS ============ */
.section-label {
    display: inline-flex;
    align-items: center;
    gap: 10px;
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.35em;
    text-transform: uppercase;
    color: var(--primary);
    margin: 3rem 0 1.2rem 0;
    padding-left: 2px;
    opacity: 0.8;
}
.section-label::after {
    content: '';
    display: block;
    width: 120px;
    height: 1px;
    background: linear-gradient(90deg, rgba(0,242,255,0.5), transparent);
}

/* ============ PROGRESS BAR ============ */
.stProgress > div > div > div > div {
    background: linear-gradient(90deg, var(--primary), var(--secondary)) !important;
    border-radius: 99px !important;
    box-shadow: 0 0 12px rgba(0,242,255,0.45) !important;
}

/* ============ DATAFRAMES & TABLES ============ */
.stDataFrame, [data-testid="stTable"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    overflow: hidden !important;
}
/* Force internal grid lines */
.stDataFrame [data-testid="styled-table-container"] {
    border: 1px solid var(--border) !important;
}
table {
    width: 100% !important;
    border-collapse: collapse !important;
}
th, td {
    border: 1px solid var(--border) !important;
    padding: 12px !important;
    text-align: left !important;
}
th {
    background: #F4B084 !important; /* Orange, Accent 6, Lighter 40% */
    color: #1a1a1a !important;     /* High contrast dark text */
    font-weight: 800 !important;
    text-transform: uppercase;
    font-size: 0.75rem;
    letter-spacing: 0.05em;
    border: 1px solid rgba(0,0,0,0.1) !important;
}

/* ============ SELECTBOX / RADIO ============ */
.stSelectbox > div > div {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    color: #fff !important;
}

/* ============ FILE UPLOADER ============ */
[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.02) !important;
    border: 1px dashed rgba(255,255,255,0.15) !important;
    border-radius: 16px !important;
    padding: 1.5rem !important;
    transition: all 0.3s ease !important;
}
[data-testid="stFileUploader"]:hover {
    border-color: var(--primary) !important;
    background: rgba(255,255,255,0.04) !important;
}
/* Fix overlapping text/labels in uploader */
[data-testid="stFileUploader"] section {
    background: transparent !important;
}
[data-testid="stFileUploader"] label {
    font-weight: 600 !important;
    margin-bottom: 0.8rem !important;
    color: var(--primary) !important;
}
/* Ensure the button inside doesn't inherit module card styles */
[data-testid="stFileUploader"] button {
    height: auto !important;
    width: auto !important;
    padding: 0.5rem 1rem !important;
    background: var(--primary) !important;
    color: #000 !important;
}

/* ============ METRICS ============ */
[data-testid="stMetricValue"] {
    color: var(--primary) !important;
    font-family: 'Outfit', sans-serif !important;
    font-weight: 800 !important;
    font-size: 2rem !important;
}

/* ============ FADE-IN ANIMATION ============ */
@keyframes fadeUp {
    from { opacity: 0; transform: translateY(18px); }
    to   { opacity: 1; transform: translateY(0); }
}
.dashboard-grid {
    animation: fadeUp 0.6s ease-out both;
}
</style>
""", unsafe_allow_html=True)



# Path to the login log file
LOGIN_LOG_FILE = "login_logs.csv"

def log_login(username):
    log_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_entry = pd.DataFrame([[username, log_time]], columns=["Username", "LoginTime"])
    if os.path.exists(LOGIN_LOG_FILE):
        new_entry.to_csv(LOGIN_LOG_FILE, mode='a', header=False, index=False)
    else:
        new_entry.to_csv(LOGIN_LOG_FILE, index=False)

# ------------------ LOGIN ------------------
# Easy-to-add users system
users = {
    "admin": "1234",
    "gokul": "abcd",
    "vel":"1234",
    "yogiram":"yogiram",
    "siva":"1234",
    "rajan":"1234",
    "bhuvana":"1234",
    "kiruba":"1234",
    "kaviya":"1234",
    "sneha":"1234"
}

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    col_l, col_c, col_r = st.columns([1, 1.5, 1])
    with col_c:
        if os.path.exists("logo.png"):
            st.image("logo.png", width=120)
        st.markdown("""
            <h1 style='font-size:2.2rem; margin-top:0.8rem; -webkit-text-fill-color:#fff;'>
                🔐 Authentication
            </h1>
            <p style='color:rgba(255,255,255,0.5); margin-bottom:1.5rem; font-size:0.95rem;'>
                Yogiram Strategic Operations Portal
            </p>
        """, unsafe_allow_html=True)
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        if st.button("Access Portal"):
            if username in users and password == users[username]:
                st.session_state.logged_in = True
                st.session_state.user = username
                st.success("Access Granted.")
                log_login(username)
                st.rerun()
            else:
                st.error("Invalid credentials.")
    st.stop()

# Get username from session
username = st.session_state.user

# ------------------ HELPER FOR LOADING PROGRESS ------------------
def load_data_with_progress(file):
    import time
    progress_bar = st.progress(0)
    for percent_complete in range(100):
        time.sleep(0.001) # fast fake progress for smooth feel
        progress_bar.progress(percent_complete + 1)
    
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    else:
        return pd.read_excel(file)

# ------------------ NAVIGATION HELPERS ------------------
def go_home(): st.session_state.page = "home"





# ------------------ NAVIGATION ------------------
if "page" not in st.session_state:
    st.session_state.page = "home"

def go_home(): st.session_state.page = "home"
def go_claim(): st.session_state.page = "claim"
def go_maxmin(): st.session_state.page = "maxmin"
def go_sales(): st.session_state.page = "sales"
def go_admin_log(): st.session_state.page = "admin_log"
def go_apollo():   # <-- New function for Apollo Check
    st.session_state.page = "apollo"
def go_pending_indents():
    st.session_state.page = "pending_indents"
def go_na_finder():
    st.session_state.page = "na_finder"
def go_info():
    st.session_state.page = "info"

def go_courier_mapper():
    st.session_state.page = "courier_mapper"
def go_pending_lock():
    st.session_state.page = "pending_lock_analyzer"
def go_aging_analysis():
    st.session_state.page = "aging_analysis"
def go_forecast():
    st.session_state.page = "pharma_forecast"

# ================== HOME PAGE ==================
if st.session_state.page == "home":

    st.markdown("""
        <div style="text-align:center; padding:3rem 0 1rem 0;">
            <div style="display:inline-block; background:linear-gradient(90deg, rgba(79, 70, 229, 0.15), rgba(6, 182, 212, 0.15));
                        border:1px solid rgba(255,255,255,0.1); border-radius:99px;
                        padding:6px 24px; font-size:0.75rem; font-weight:800;
                        letter-spacing:0.35em; color:#fff; text-transform:uppercase;
                        margin-bottom:1.5rem; box-shadow: 0 0 20px rgba(6, 182, 212, 0.2);">
                ⚡ COMMAND CENTER
            </div>
            <h1 style="font-size:clamp(2.5rem,6vw,4rem); font-weight:800; letter-spacing:-2px;
                       background:linear-gradient(135deg,#ffffff 0%,#a5b4fc 100%);
                       -webkit-background-clip:text; -webkit-text-fill-color:transparent;
                       margin:0 0 0.5rem 0; line-height:1.1; text-shadow: 0 4px 24px rgba(0,0,0,0.4);">
                Yogiram Automation
            </h1>
            <p style="color:#94a3b8; font-size:1.1rem; font-weight:400; margin:0; letter-spacing:0.02em;">
                Strategic Business Intelligence & Operations Portal
            </p>
        </div>
    """, unsafe_allow_html=True)



    # Frequently Used section header
    st.markdown("""
        <div style="display:flex;align-items:center;gap:18px;margin:3.5rem 0 2rem 0;">
            <span style="font-size:0.75rem;font-weight:800;letter-spacing:0.4em;
                         text-transform:uppercase;color:#e2e8f0;white-space:nowrap;">
                <span style="color:#06b6d4;">●</span> FREQUENTLY USED
            </span>
            <div style="flex:1;height:1px;background:linear-gradient(90deg,rgba(255,255,255,0.1),transparent);"></div>
        </div>
    """, unsafe_allow_html=True)

    t1, t2, t3, t4 = st.columns(4, gap="medium")
    with t1:
        if st.button("ℹ️  Info Guide",       key="btn_info", width='stretch'):   go_info()
        if st.button("📈  Pharma Forecast",  key="btn_fc", width='stretch'):     go_forecast()
    with t2:
        if st.button("🛒  Apollo Check",    key="btn_apollo", width='stretch'): go_apollo()
        if st.button("📦  Pending Indents", key="btn_pi", width='stretch'):     go_pending_indents()
    with t3:
        if st.button("🔒  Order Lock",      key="btn_ol", width='stretch'):     go_pending_lock()
        if st.button("⏳  Due List",              key="btn_due", width='stretch'):    go_aging_analysis()
    with t4:
        if st.button("🧾  NA Finder",       key="btn_na", width='stretch'):     go_na_finder()
        if st.button("💰  Sales Portal",   key="btn_sales", width='stretch'):  go_sales()

    # Additional Modules section header
    st.markdown("""
        <div style="display:flex;align-items:center;gap:18px;margin:3.5rem 0 2rem 0;">
            <span style="font-size:0.75rem;font-weight:800;letter-spacing:0.4em;
                         text-transform:uppercase;color:#e2e8f0;white-space:nowrap;">
                <span style="color:#f43f5e;">●</span> ADDITIONAL MODULES
            </span>
            <div style="flex:1;height:1px;background:linear-gradient(90deg,rgba(255,255,255,0.1),transparent);"></div>
        </div>
    """, unsafe_allow_html=True)

    a1, a2, a3, a4 = st.columns(4, gap="medium")
    with a1:
        if st.button("📂  Claim Portal",      key="btn_claim", width='stretch'):   go_claim()
        if st.button("📊  Max Min Portal",    key="btn_mm", width='stretch'):      go_maxmin()
    with a2:
        if st.button("📊  DB Age Analysis",   key="btn_db", width='stretch'):      st.session_state.page = "db_age"
    with a3:
        if st.button("💹  Contribution",      key="btn_contrib", width='stretch'): st.session_state.page = "sales_contribution"
        if st.button("🧠  AI Analyst",        key="btn_ai", width='stretch'):      st.session_state.page = "ai_data_assistant"
    with a4:
        if st.button("🚚  Courier Map",     key="btn_cm", width='stretch'):     go_courier_mapper()

    if username == "admin":
        st.markdown("""
            <div style="display:flex;align-items:center;gap:18px;margin:3.5rem 0 2rem 0;">
                <span style="font-size:0.75rem;font-weight:800;letter-spacing:0.4em;
                             text-transform:uppercase;color:#e2e8f0;white-space:nowrap;">
                    <span style="color:#f59e0b;">●</span> ADMIN CONTROLS
                </span>
                <div style="flex:1;height:1px;background:linear-gradient(90deg,rgba(255,255,255,0.1),transparent);"></div>
            </div>
        """, unsafe_allow_html=True)
        adm1, adm2, adm3, adm4 = st.columns(4, gap="medium")
        with adm1:
            if st.button("📝  Login Activity", key="btn_admin", width='stretch'): go_admin_log()



    st.markdown("""
        <div style="margin-top:6rem; padding:2rem 0; border-top:1px solid rgba(255,255,255,0.08);
                    display:flex; justify-content:space-between; align-items:flex-end;">
            <div style="color:#64748b; font-size:0.8rem; line-height:1.6;">
                <p style="margin:0; letter-spacing:0.15em; font-weight:800; color:#94a3b8;">YOGIRAM AUTOMATION</p>
                <p style="margin:0;">Signed in as <strong style="color:#06b6d4;">{user}</strong></p>
            </div>
            <div style="text-align:right;">
                <p style="margin:0; font-size:0.7rem; text-transform:uppercase; letter-spacing:0.2em; color:#64748b; font-weight:800;">Developed By</p>
                <p style="margin:0; font-size:1.15rem; font-weight:800; background:linear-gradient(135deg,#06b6d4,#4f46e5); -webkit-background-clip:text; -webkit-text-fill-color:transparent; font-family:'Outfit';">Gokul Srinivas</p>
                <p style="margin:0; font-size:0.8rem; color:#64748b; font-weight:600; letter-spacing:0.05em;">Data Analyst</p>
            </div>
        </div>
    """.format(user=username), unsafe_allow_html=True)



# ------------------ INFO MODULE ------------------
elif st.session_state.page == "info":
    st.title("ℹ️️ Info & Instructions Portal")

    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"

    st.markdown("""
    ### 📂 Claim Portal
    - Upload the claim report file (Excel/CSV).
    - Groups data by 'Company Name' and sums 'Claim Amount'.
    - Generates a company-wise total claim summary for download.

    ### 📊 Max/Min Value Portal
    - Upload any dataset with numeric columns.
    - Quickly identifies the Maximum and Minimum values across all numeric fields.

    ### 💰 Sales Portal
    - Upload a detailed Sales Export file.
    - Analyzes performance by Weekday, Company, Outlet, Product, and Salesman.
    - Provides a multi-sheet Excel summary for deep-dive analysis.

    ### 🛒 Apollo Check Portal
    - Analyzes order frequency and consistency for Apollo shops.
    - Tracks 'Continuous' order streaks (e.g., items ordered 5 days in a row).
    - Tracks 'Non-Continuous' total order counts within a period.

    ### 📦 Pending Indents Check
    - Upload 'Pending Indents' and 'Order Details' files.
    - Maps ContractIDs and calculates Fulfillment % (NS %).
    - Generates an aggregated summary by ContractID.

    ### 🧾 NA Finder
    - Upload 'Indent Data' and 'Item Master'.
    - Aggregates NA/NS counts and values by SKU/Company.
    - Merges current stock quantity from the Item Master for a complete status report.

    ### 📊 DB Age Analysis
    - Calculates the "Age" of supplier claims based on a reference date.
    - Buckets claims into 0-30, 31-60, 61-90... Above 360 days.
    - Generates a Supplier-wise pivot table of pending amounts.

    ### 💹 Sales Contribution Analyzer
    - Performs Pareto analysis (80/20 rule) on sales data.
    - Rank products by sales amount and identify the Top X% or Bottom X% contributors.
    - Features "Nested Selection" to drill down into high-performing subsets.

    ### 📦 Courier Bill Count Portal
    - Merges Transaction files, Courier details, and Tray details.
    - Groups bills by Account No. and adds specific Tray IDs for courier reporting.
    - Generates a formatted report ready for courier dispatch.

    ### 🧠 AI Data Analyst Portal
    - Interactive AI assistant powered by Groq (Llama 3).
    - Upload any data and ask questions in plain English (e.g., "What is my top company by profit?").
    - AI automatically writes code, generates KPIs, and creates Plotly charts.

    ### 📦 Pending Order Lock Analyzer
    - Maps pending orders to "Lock Reasons" and "Salesmen".
    - Helps identify why specific orders are stuck and who is responsible for them.

    ### ⏳ Due List Checker
    - Upload raw dues data with a **'DAYS'** numeric column.
    - System automatically creates buckets: 0-30, 30-60, 60-90, 90-120, 120-180, 180-360, and Above 360 days.
    - Provides a searchable multi-select tool to download individual "Dues Lists" for specific salesmen.

    ---
    **General Tips:**
    - Always ensure your files are in .xlsx or .csv format.
    - Match column names as requested in each module.
    - Use the 'Back to Home' button to switch between different tools.
    """)

# ------------------ AI DATA ASSISTANT (ADVANCED) ------------------
elif st.session_state.page == "ai_data_assistant":

    import pandas as pd
    import streamlit as st
    from groq import Groq
    from io import BytesIO
    import plotly.express as px

    st.title("🧠 AI Data Analyst (Smart + Interactive)")

    # ---------------- BACK ----------------
    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"

    # ---------------- GROQ ----------------
    client = Groq(api_key="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")

    # ---------------- FILE UPLOAD ----------------
    uploaded = st.file_uploader("📂 Upload Excel / CSV", type=["xlsx", "xls", "csv"])

    df = None
    if uploaded:
        df = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
        st.success("✅ File loaded")
        st.dataframe(df.head(), width='stretch')

    # ---------------- EXCEL DOWNLOAD ----------------
    def download_excel(dataframe):
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            dataframe.to_excel(writer, index=False)
        buffer.seek(0)
        return buffer

    # ---------------- ASK AI ----------------
    if df is not None:
        query = st.text_input("💬 Ask about your data (KPIs, charts, insights, summaries)")

        if query:
            prompt = f"""
You are a SENIOR BUSINESS DATA ANALYST.

Dataframe name: df
Columns: {list(df.columns)}

User question:
{query}

YOUR RESPONSIBILITIES:
1. Understand business intent
2. Generate:
   - KPI(s)
   - Interactive chart (Plotly)
   - Short business summary
3. Ask clarification ONLY if mandatory

OUTPUT RULES (VERY STRICT):
- Output ONLY python code
- Use plotly.express for charts
- Use streamlit functions
- Do NOT explain code
- No markdown text except st.write / st.markdown

KPI FORMAT:
st.metric("Title", value)

SUMMARY FORMAT:
st.subheader("📌 Summary")
st.write("Natural language business explanation")

TABLE FORMAT:
result_df = ...
st.dataframe(result_df, width='stretch')

CHART FORMAT:
fig = px.bar / px.line / px.pie
st.plotly_chart(fig, width='stretch')

SUMMARY INTELLIGENCE RULES:
- Calculate contribution %
- Mention top contributor %
- Mention concentration (top 5 / top 10 share)

CLARIFICATION RULE:
If chart type unclear:
st.warning("Which chart do you want? (bar / line / pie)")
STOP execution after warning
"""

            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0
            )

            ai_code = response.choices[0].message.content
            ai_code = ai_code.replace("```python", "").replace("```", "").strip()

            st.code(ai_code, language="python")

            try:
                env = {
                    "df": df,
                    "pd": pd,
                    "st": st,
                    "px": px
                }
                exec(ai_code, env)

                if "result_df" in env and isinstance(env["result_df"], pd.DataFrame):
                    st.download_button(
                        "⬇️ Download Result as Excel",
                        data=download_excel(env["result_df"]),
                        file_name="AI_Result.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            except Exception as e:
                st.error("⚠️ AI execution error")
                st.exception(e)


# ------------------ CLAIM PORTAL ------------------
elif st.session_state.page == "claim":
    st.title("📂 Free Claim Portal")
    if st.button("🏠 Back to Home"): go_home()

    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "csv"])
    if uploaded_file:
        df = pd.read_excel(uploaded_file) if uploaded_file.name.endswith(".xlsx") else pd.read_csv(uploaded_file)
        st.write("### Uploaded Data Preview:")
        st.dataframe(df.head())

        if st.button("🚀 Process Data"):
            grouped = df.groupby("Company Name")["Claim Amount"].sum().reset_index()
            grouped.columns = ["Company Name", "Total Claim Amount"]
            st.write("### 📊 Results Summary")
            st.write(f"**Total Claims:** {len(df)}")
            st.write("**Top 5 Companies by Claim Amount:**")
            st.dataframe(grouped.sort_values("Total Claim Amount", ascending=False).head(5))

            def to_excel(df):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Summary')
                return output.getvalue()

            excel_data = save_df_to_excel_with_format(grouped, sheet_name="Summary")
            today_date = datetime.now().strftime("%Y-%m-%d")
            file_name = f"Company_Wise_Free_Claim_Issued_{today_date}.xlsx"
            st.download_button("📥 Download Processed File", data=excel_data, file_name=file_name,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ------------------ MAX-MIN PORTAL ------------------
elif st.session_state.page == "maxmin":
    st.title("📊 Max-Min Value Portal")
    if st.button("🏠 Back to Home"): go_home()

    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "csv"])
    if uploaded_file:
        df = pd.read_excel(uploaded_file) if uploaded_file.name.endswith(".xlsx") else pd.read_csv(uploaded_file)
        st.write("### Uploaded Data Preview:")
        st.dataframe(df.head())

        if st.button("🔍 Find Max and Min Values"):
            numeric_df = df.select_dtypes(include='number')
            st.write("### 🔺 Maximum Values:")
            st.dataframe(numeric_df.max().to_frame("Max Value"))
            st.write("### 🔻 Minimum Values:")
            st.dataframe(numeric_df.min().to_frame("Min Value"))

# ------------------ SALES PORTAL ------------------
elif st.session_state.page == "sales":
    st.title("💰 Sales Analysis Portal")
    if st.button("🏠 Back to Home"): go_home()

    uploaded_file = st.file_uploader("Upload your Sales Excel/CSV file", type=["xlsx", "csv"])
    if uploaded_file:
        # Suppress DtypeWarning with low_memory=False
        if uploaded_file.name.endswith(".xlsx"):
            df = pd.read_excel(uploaded_file)
        else:
            df = pd.read_csv(uploaded_file, low_memory=False)
            
        st.write("### Uploaded Data Preview:")
        st.dataframe(df.head())

        if st.button("📊 Analyze Sales"):
            # Normalize column detection
            def get_col(candidates):
                for c in candidates:
                    for col in df.columns:
                        if col.strip().lower() == c.lower():
                            return col
                return None

            dated_col = get_col(["Dated", "Date", "Bill Date"])
            value_col = get_col(["Value", "Amount", "Amt", "Bill Amt", "Bill Amount"])
            company_col = get_col(["Company", "Party", "Vendor", "Company Name"])
            outlet_col = get_col(["Outlet", "Branch", "Store", "Outlet Name"])
            product_col = get_col(["Product", "Item Name", "Item", "Product Name"])
            salesman_col = get_col(["SalesMan", "Salesman", "Sales Man", "Sales Person", "SalesMan Name"])

            # Check required columns
            missing = []
            if not dated_col: missing.append("Dated/Date")
            if not value_col: missing.append("Value/Amount")
            if missing:
                st.error(f"⚠️ Missing required columns: {', '.join(missing)}")
                st.stop()

            # Process data
            df[dated_col] = pd.to_datetime(df[dated_col], errors='coerce', dayfirst=True)
            df['Weekday'] = df[dated_col].dt.day_name()
            df[value_col] = pd.to_numeric(df[value_col], errors='coerce').fillna(0)

            total_sales = df[value_col].sum()
            st.metric("💵 Total Sales", f"{total_sales:,.2f}")

            # 1. Sales by Weekday
            sales_by_day = df.groupby('Weekday')[value_col].sum().reindex(
                ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']).reset_index()
            st.write("### 🗓 Total Sales by Day")
            st.dataframe(sales_by_day, width='stretch')

            # 2. Sales by Company
            if company_col:
                company_sales = df.groupby(company_col)[value_col].sum().reset_index()
                st.write(f"### 🏢 Total Sales by {company_col}")
                st.dataframe(company_sales.sort_values(value_col, ascending=False), width='stretch')
            
            # 3. Sales by Outlet
            if outlet_col:
                outlet_sales = df.groupby(outlet_col)[value_col].sum().reset_index()
                st.write(f"### 🏪 Total Sales by {outlet_col}")
                st.dataframe(outlet_sales.sort_values(value_col, ascending=False), width='stretch')

            # 4. Sales by Product
            if product_col:
                product_sales = df.groupby(product_col)[value_col].sum().reset_index()
                st.write(f"### 📦 Total Sales by {product_col}")
                st.dataframe(product_sales.sort_values(value_col, ascending=False), width='stretch')

            # 5. Sales by Salesman
            if salesman_col:
                salesman_sales = df.groupby(salesman_col)[value_col].sum().reset_index()
                st.write(f"### 👨‍💼 Total Sales by {salesman_col}")
                st.dataframe(salesman_sales.sort_values(value_col, ascending=False), width='stretch')

            # Generate formatted Excel export
            export_dict = {"Sales_by_Day": sales_by_day}
            if company_col: export_dict[f"Sales_by_{company_col[:20]}"] = company_sales
            if outlet_col: export_dict[f"Sales_by_{outlet_col[:20]}"] = outlet_sales
            if product_col: export_dict[f"Sales_by_{product_col[:20]}"] = product_sales
            if salesman_col: export_dict[f"Sales_by_{salesman_col[:20]}"] = salesman_sales

            excel_data = save_multi_df_to_excel_with_format(export_dict)

            st.download_button("📥 Download Sales Summary (Formatted)", data=excel_data,
                               file_name=f"Sales_Analysis_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# ------------------ PENDING INDENTS MODULE ------------------
elif st.session_state.page == "pending_indents":
    st.title("📦 Pending Indents Module")
    
    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"

    # Upload files
    pending_file = st.file_uploader("Upload Pending Indents File", type=["xlsx","csv"], key="pending")
    second_file = st.file_uploader("Upload Order Details File", type=["xlsx","csv"], key="second")

    if pending_file and second_file:
        df_pending = pd.read_excel(pending_file) if pending_file.name.endswith(".xlsx") else pd.read_csv(pending_file)
        df_second = pd.read_excel(second_file) if second_file.name.endswith(".xlsx") else pd.read_csv(second_file)

        st.write("### Uploaded Files Preview")
        st.write("Pending Indents:")
        st.dataframe(df_pending.head())
        st.write("Order Details:")
        st.dataframe(df_second.head())

        if st.button("🚀 Map & Aggregate Data"):
            # Map ContractID from Pending Indents
            mapping = df_pending.set_index("Ind.No.")["ContractID"].to_dict()
            df_second["ContractID"] = df_second["Ind No"].map(mapping).fillna("Manual")
            df_second["ContractID"] = df_second["ContractID"].replace("", "(blank)")

            # Aggregate and calculate Fulfillment %
            agg_df = df_second.groupby("ContractID").agg({"Ordered Items":"sum", "Invoice Items":"sum"}).reset_index()
            # Fulfillment % as proper percentage format
            agg_df["Fulfillment %"] = ((agg_df["Invoice Items"] / agg_df["Ordered Items"]) * 100).round(2).astype(str) + "%"


            # Grand Total row
            grand_total = pd.DataFrame({
                "ContractID": ["Grand Total"],
                "Ordered Items": [agg_df["Ordered Items"].sum()],
                "Invoice Items": [agg_df["Invoice Items"].sum()],
                "Fulfillment %": [(agg_df["Invoice Items"].sum() / agg_df["Ordered Items"].sum() * 100).round(2)]
            })

            final_df = pd.concat([agg_df, grand_total], ignore_index=True)
            st.write("### 📊 Aggregated Results")
            st.dataframe(final_df)

            # Download Excel
            from io import BytesIO
            def to_excel(df):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Summary')
                return output.getvalue()

            st.download_button("📥 Download Excel", data=save_df_to_excel_with_format(final_df, sheet_name="Summary"), file_name="Pending_Indents_Summary.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # Download CSV
            st.download_button("📥 Download CSV", data=final_df.to_csv(index=False).encode('utf-8'), 
                               file_name="Pending_Indents_Summary.csv", mime="text/csv")



# ------------------ NA FINDER MODULE ------------------
elif st.session_state.page == "na_finder":
    st.title("🧮 NA Finder Module")
    if st.button("🏠 Back to Home"): go_home()

    # Upload files
    file1 = st.file_uploader("Upload first file (Indent Data)", type=["xlsx","csv"], key="file1")
    file2 = st.file_uploader("Upload second file (Item Master)", type=["xlsx","csv"], key="file2")

    if file1 and file2:
        try:
            df1 = pd.read_excel(file1) if file1.name.endswith(".xlsx") else pd.read_csv(file1)
            df2 = pd.read_excel(file2) if file2.name.endswith(".xlsx") else pd.read_csv(file2)

            # Normalize column names
            df1.columns = [str(c).strip() for c in df1.columns]
            df2.columns = [str(c).strip() for c in df2.columns]

            # Safely cast all non-numeric columns to string to avoid Arrow/PyArrow crashes
            for col in df1.columns:
                if df1[col].dtype == object:
                    df1[col] = df1[col].fillna("").astype(str)
            for col in df2.columns:
                if df2[col].dtype == object:
                    df2[col] = df2[col].fillna("").astype(str)

            st.write("### File Previews")
            st.write("Indent Data:")
            st.dataframe(df1.head())
            st.write("Item Master Data:")
            st.dataframe(df2.head())

        except Exception as e:
            st.error(f"⚠️ Error reading files: {e}")
            st.stop()

        # ---------- Process Button ----------
        if st.button("🚀 Process & Merge Qty from Item Master"):
            try:
                # Check required columns in df1
                required_df1 = ['SKU CODE', 'CODE', 'ITEM NAME', 'COMPANY', 'NA', 'NA VALUE']
                missing_df1 = [c for c in required_df1 if c not in df1.columns]
                if missing_df1:
                    st.error(f"⚠️ Indent Data is missing columns: {', '.join(missing_df1)}")
                    st.stop()

                # Convert numeric cols back from str
                for col in ['NA', 'NA VALUE']:
                    df1[col] = pd.to_numeric(df1[col], errors='coerce').fillna(0)

                grouped_df = df1.groupby(
                    ['SKU CODE', 'CODE', 'ITEM NAME', 'COMPANY'],
                    as_index=False
                ).agg({
                    'NA': ['count', 'sum'],
                    'NA VALUE': 'sum'
                })

                grouped_df.columns = [
                    'SKU CODE', 'Gold Code', 'ITEM NAME', 'COMPANY',
                    'Count of NA', 'Sum of NA', 'Sum of NA VALUE'
                ]

                if "Gold Code" not in df2.columns or "Qty" not in df2.columns:
                    st.error("⚠️ Item Master file must have 'Gold Code' and 'Qty' columns")
                    st.stop()

                grouped_df['Gold Code'] = grouped_df['Gold Code'].astype(str).str.strip()
                df2['Gold Code'] = df2['Gold Code'].astype(str).str.strip()

                merged_df = grouped_df.merge(
                    df2[['Gold Code', 'Qty']],
                    on='Gold Code',
                    how='left'
                )

                merged_df['Qty'] = pd.to_numeric(merged_df['Qty'], errors='coerce').fillna(0)

                merged_df = merged_df[['SKU CODE', 'Gold Code', 'ITEM NAME', 'COMPANY', 'Qty',
                                       'Count of NA', 'Sum of NA', 'Sum of NA VALUE']]

                # Store in session_state so it persists after text input
                st.session_state["merged_df"] = merged_df
                st.success("✅ Processing completed! Scroll down for more options.")

            except Exception as e:
                st.error(f"⚠️ Processing error: {e}")

        # ---------- If processed data exists ----------
        if "merged_df" in st.session_state:
            merged_df = st.session_state["merged_df"]

            st.write("### ✅ Final Merged Result with Qty from Item Master")
            st.dataframe(merged_df)

            # ---------- Download buttons ----------
            excel_data = save_df_to_excel_with_format(merged_df, sheet_name="Merged")
            st.download_button("📥 Download Final Excel", data=excel_data,
                               file_name="NA_Finder_Final.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            st.download_button("📥 Download Final CSV", data=merged_df.to_csv(index=False).encode('utf-8'),
                               file_name="NA_Finder_Final.csv",
                               mime="text/csv")

            # ---------- New Section: Percentage Calculation ----------
            st.markdown("---")
            st.subheader("📊 Company-wise NA Percentage Calculator")

            user_input = st.text_input("Enter a number for percentage calculation (e.g., total value):")

            if user_input:
                try:
                    base_value = float(user_input)

                    # Group by company and calculate sum of Count of NA
                    company_summary = merged_df.groupby('COMPANY', as_index=False)['Count of NA'].sum()

                    # Calculate percentage
                    company_summary['Percentage'] = ((company_summary['Count of NA'] / base_value) * 100).round(3)
                    company_summary['Percentage'] = company_summary['Percentage'].astype(str) + " %"

                    st.write("### 📈 Company-wise NA Summary")
                    st.dataframe(company_summary)

                    # Optional download
                    excel_summary = save_df_to_excel_with_format(company_summary, sheet_name="Summary")
                    st.download_button("📥 Download Company Summary Excel", data=excel_summary,
                                       file_name="Company_NA_Summary.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                except ValueError:
                    st.error("⚠️ Please enter a valid numeric value.")




# ------------------ DB AGE ANALYSIS MODULE ------------------
elif st.session_state.page == "db_age":
    st.title("📊 DB Age Analysis Module")
    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"

    # ------------------ FILE UPLOADER ------------------
    uploaded_file = st.file_uploader("Upload your Excel/CSV file", type=["xlsx","csv"], key="db_age_file")
    if uploaded_file:
        df = pd.read_excel(uploaded_file) if uploaded_file.name.endswith(".xlsx") else pd.read_csv(uploaded_file)
        
        # Strip spaces from column names
        df.columns = df.columns.str.strip()
        
        # Find DB Date column
        db_date_col = [col for col in df.columns if col.lower() == 'db date']
        if not db_date_col:
            st.error("⚠️ No 'DB Date' column found in the uploaded file.")
            st.stop()
        db_date_col = db_date_col[0]

        st.write("### Uploaded Data Preview")
        st.dataframe(df.head())

        # ------------------ DATE SELECTION ------------------
        st.subheader("Select Reference Date for Age Calculation")
        date_option = st.radio("Choose date option:", ["Today", "Custom Date"])
        if date_option == "Custom Date":
            reference_date = st.date_input("Select Date")
        else:
            reference_date = pd.to_datetime("today")

        # ------------------ AGE CALCULATION ------------------
        df['DB Date'] = pd.to_datetime(df[db_date_col], errors='coerce', dayfirst=True)
        df['AgeDays'] = (pd.to_datetime(reference_date) - df['DB Date']).dt.days

        # ------------------ AGE BUCKETS ------------------
        def age_bucket(days):
            if pd.isna(days):
                return "Unknown"
            if 0 <= days <= 30:
                return "0-30"
            elif 31 <= days <= 60:
                return "31-60"
            elif 61 <= days <= 90:
                return "61-90"
            elif 91 <= days <= 120:
                return "91-120"
            elif 121 <= days <= 180:
                return "121-180"
            elif 181 <= days <= 270:
                return "181-270"
            elif 271 <= days <= 360:
                return "271-360"
            else:
                return "Above 360"

        df['Age Bucket'] = df['AgeDays'].apply(age_bucket)

        st.write("### Data with Age and Buckets")
        st.dataframe(df.head())

        # ------------------ PIVOT TABLE ------------------
        if 'Pending' not in df.columns:
            st.error("⚠️ No 'Pending' column found in the uploaded file.")
        else:
            pivot_df = pd.pivot_table(
                df,
                index='Supplier',
                columns='Age Bucket',
                values='Pending',
                aggfunc='sum',
                fill_value=0
            ).reset_index()

            st.write("### 📊 Supplier-wise Pending Amount by Age Bucket")
            st.dataframe(pivot_df)

            # ------------------ DOWNLOAD ------------------
            from io import BytesIO


            excel_data = save_df_to_excel_with_format(pivot_df, sheet_name="DB_Age_Analysis")
            st.download_button("📥 Download DB Age Pivot Excel", data=excel_data,
                               file_name=f"DB_Age_Analysis_{pd.to_datetime(reference_date).strftime('%Y-%m-%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ------------------ SALES CONTRIBUTION ANALYZER ------------------
elif st.session_state.page == "sales_contribution":
    import pandas as pd
    import streamlit as st
    from io import BytesIO

    st.title("📈 Sales Contribution Analyzer")

    if st.button("🏠 Back to Home"):
        go_home()

    st.markdown("Upload your sales/stock export (CSV or Excel). Expected columns include `AMOUNT` and `QTY.` (common variants handled).")

    # ---------- Upload & cache ----------
    uploaded = st.file_uploader("Upload sales file (CSV / XLSX)", type=["csv", "xlsx"], key="sales_contrib_upload")
    if uploaded:
        # store in session_state to avoid repeated reads
        if "sales_raw_df" not in st.session_state or st.session_state.get("sales_file_name") != uploaded.name:
            try:
                if uploaded.name.lower().endswith(".csv"):
                    df_raw = pd.read_csv(uploaded)
                else:
                    df_raw = pd.read_excel(uploaded)
            except Exception as e:
                st.error(f"Error reading file: {e}")
                st.stop()

            st.session_state.sales_raw_df = df_raw.copy()
            st.session_state.sales_file_name = uploaded.name

    # allow working with previously uploaded file if present
    if "sales_raw_df" not in st.session_state:
        st.info("No file uploaded yet.")
        st.stop()

    df_raw = st.session_state.sales_raw_df.copy()

    # ---------- Normalize column names and detect key columns ----------
    df_raw.columns = df_raw.columns.map(lambda c: str(c).strip())

    # Helper to find column by possible names
    def find_col(df, candidates):
        cols = df.columns.tolist()
        for c in candidates:
            for col in cols:
                if col.lower().strip() == c.lower().strip():
                    return col
        # fuzzy-like match: remove dots and spaces and compare
        normalized = { "".join(ch.lower() for ch in col if ch.isalnum()): col for col in cols }
        for c in candidates:
            key = "".join(ch.lower() for ch in c if ch.isalnum())
            if key in normalized:
                return normalized[key]
        return None

    amount_col = find_col(df_raw, ["AMOUNT", "AMT", "Amount", "Amount "])
    qty_col = find_col(df_raw, ["QTY.", "QTY", "QTY", "Quantity", "QTY", "Qty", "QTY"])
    item_col = find_col(df_raw, ["ITEM NAME", "ITEM", "ITEMNAME", "BARCODE", "BARCODE"])
    barcode_col = find_col(df_raw, ["BARCODE", "BAR CODE", "BARCODE "])
    company_col = find_col(df_raw, ["COMPANY", "COMPANY "])

    # Validate
    if amount_col is None:
        st.error("⚠️ Could not find an 'Amount' column. Expected column names like 'AMOUNT' or 'AMT'.")
        st.stop()
    if qty_col is None:
        st.error("⚠️ Could not find a 'Qty' column. Expected 'QTY.' or 'QTY'.")
        st.stop()
    if item_col is None and barcode_col is None:
        st.error("⚠️ Could not find an 'Item Name' or 'Barcode' column. At least one is required.")
        st.stop()

    # Normalize working df
    df = df_raw.copy()
    df.rename(columns={amount_col: "amount", qty_col: "qty"}, inplace=True)
    if item_col:
        df.rename(columns={item_col: "item_name"}, inplace=True)
    if barcode_col:
        df.rename(columns={barcode_col: "barcode"}, inplace=True)
    if company_col:
        df.rename(columns={company_col: "company"}, inplace=True)

    # Ensure numeric types
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0)

    # ---------- Overall KPIs ----------
    total_sales_all = df["amount"].sum()
    total_qty_all = df["qty"].sum()
    # define product identifier: barcode if exists else item_name
    if "barcode" in df.columns:
        df["_prod_id"] = df["barcode"].astype(str).str.strip()
        prod_label = "barcode"
    else:
        df["_prod_id"] = df["item_name"].astype(str).str.strip()
        prod_label = "item_name"

    total_products_all = df["_prod_id"].nunique()

    st.markdown("### 🔢 Overall KPIs")
    k1, k2, k3 = st.columns(3)
    k1.metric("💵 Total Sales (All)", f"₹{total_sales_all:,.2f}")
    k2.metric("🔢 Total Products", f"{total_products_all:,}")
    k3.metric("📦 Total Quantity", f"{total_qty_all:,.0f}")

    # ---------- Group and Rank ----------
    # Group by product id and show item_name & company where available
    group_cols = ["_prod_id"]
    agg = df.groupby(group_cols).agg({
        "amount": "sum",
        "qty": "sum"
    }).reset_index()

    # attach item_name and company (first occurrence)
    if "item_name" in df.columns:
        first_item = df.groupby("_prod_id")["item_name"].first().reset_index()
        agg = agg.merge(first_item, on="_prod_id", how="left")
    if "company" in df.columns:
        first_comp = df.groupby("_prod_id")["company"].first().reset_index()
        agg = agg.merge(first_comp, on="_prod_id", how="left")

    agg = agg.sort_values("amount", ascending=False).reset_index(drop=True)
    agg["rank"] = agg.index + 1
    agg["cum_amount"] = agg["amount"].cumsum()
    agg["cum_pct"] = (agg["cum_amount"] / agg["amount"].sum() * 100).round(4)

    # friendly display order
    display_cols = ["rank", "_prod_id"]
    if "item_name" in agg.columns:
        display_cols.append("item_name")
    if "company" in agg.columns:
        display_cols.append("company")
    display_cols += ["amount", "qty", "cum_amount", "cum_pct"]

    st.markdown("### 📋 Ranked Products (by Sales)")
    st.dataframe(agg[display_cols].rename(columns={"_prod_id": prod_label}), width='stretch')

    # ---------- Percentage selection ----------
    st.markdown("---")
    st.subheader("Select contribution filter")

    col1, col2 = st.columns([2,3])
    mode = col1.selectbox("Choose mode", ["Top % by Sales", "Bottom % by Sales"])
    percent = col2.slider("Select percentage (X%)", min_value=1, max_value=100, value=80, step=1)

    # function to pick minimal set of products reaching >= pct of sales
    def pick_by_percent(df_ranked, pct, top=True):
        df_local = df_ranked.copy().reset_index(drop=True)
        total = df_local["amount"].sum()
        if total == 0:
            return df_local.iloc[0:0]  # empty
        if top:
            df_local = df_local.sort_values("amount", ascending=False).reset_index(drop=True)
        else:
            df_local = df_local.sort_values("amount", ascending=True).reset_index(drop=True)
        df_local["cum"] = df_local["amount"].cumsum()
        cutoff = total * (pct / 100.0)
        # find first index where cum >= cutoff
        idx = df_local[df_local["cum"] >= cutoff].index
        if len(idx) == 0:
            # not reached: return all
            return df_local
        last_idx = idx[0]
        return df_local.loc[:last_idx].drop(columns=["cum"])

    top_level_selected = pick_by_percent(agg, percent, top=(mode == "Top % by Sales"))
    # compute KPIs for this selection
    sel_sales = top_level_selected["amount"].sum()
    sel_qty = top_level_selected["qty"].sum()
    sel_count = len(top_level_selected)
    sel_pct_of_total = (sel_sales / total_sales_all * 100) if total_sales_all else 0

    # ---------- Nested percent option ----------
    st.markdown("### Nested selection (optional)")
    nested_enabled = st.checkbox("Apply another % inside the selected set (e.g., top 50% of the top 80%)", value=False)
    nested_df = top_level_selected.copy()
    nested_percent = None
    if nested_enabled:
        nested_percent = st.slider("Nested percentage (Y%)", min_value=1, max_value=100, value=50, step=1, key="nested_pct")
        # For nested, we must pick within the selected set by amount proportion of that subset
        if len(nested_df) > 0:
            nested_df = pick_by_percent(nested_df, nested_percent, top=(mode == "Top % by Sales"))
        nested_sel_sales = nested_df["amount"].sum()
        nested_sel_qty = nested_df["qty"].sum()
        nested_sel_count = len(nested_df)
        nested_sel_pct_of_total = (nested_sel_sales / total_sales_all * 100) if total_sales_all else 0
    else:
        nested_sel_sales = nested_sel_qty = nested_sel_count = nested_sel_pct_of_total = None

    # ---------- Show selection KPIs ----------
    st.markdown("### ✅ Selected Set KPIs")
    s1, s2, s3, s4 = st.columns(4)
    s1.metric("Selected Sales (₹)", f"{sel_sales:,.2f}")
    s2.metric("Selected Qty", f"{sel_qty:,.0f}")
    s3.metric("No. Products Selected", f"{sel_count}")
    s4.metric("% of Total Sales", f"{sel_pct_of_total:.2f}%")

    if nested_enabled:
        st.markdown("### ✅ Nested Selected KPIs")
        n1, n2, n3, n4 = st.columns(4)
        n1.metric("Nested Sales (₹)", f"{nested_sel_sales:,.2f}")
        n2.metric("Nested Qty", f"{nested_sel_qty:,.0f}")
        n3.metric("No. Products Nested", f"{nested_sel_count}")
        n4.metric("% of Total Sales (Nested)", f"{nested_sel_pct_of_total:.2f}%")

    # ---------- Display selected tables ----------
    st.markdown(f"### 🔎 {mode} — {percent}% selection (Products: {sel_count})")
    sel_display = top_level_selected[display_cols].rename(columns={"_prod_id": prod_label}).sort_values("amount", ascending=False)
    st.dataframe(sel_display, width='stretch')

    if nested_enabled:
        st.markdown(f"### 🔎 Nested selection ({nested_percent}%) inside the above (Products: {len(nested_df)})")
        nested_display = nested_df[display_cols].rename(columns={"_prod_id": prod_label}).sort_values("amount", ascending=False)
        st.dataframe(nested_display, width='stretch')

    # ---------- Download options ----------


    col_down_1, col_down_2 = st.columns(2)
    with col_down_1:
        st.download_button(
            "📥 Download Ranked Full (Excel)",
            data=to_excel_bytes(agg[display_cols].rename(columns={"_prod_id": prod_label})),
            file_name=f"ranked_products_{pd.Timestamp.now().date()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    with col_down_2:
        # create a combined excel if nested, else selected excel
        if nested_enabled:
            # create multi-sheet workbook
            out = BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                sel_display.to_excel(writer, index=False, sheet_name="Selected_Summary")
                nested_display.to_excel(writer, index=False, sheet_name=f"Nested_{nested_percent}pct")
            data_bytes = out.getvalue()
            st.download_button(
                "📥 Download Selected + Nested (Excel)",
                data=save_multi_df_to_excel_with_format({"Selected_Summary": sel_display, f"Nested_{nested_percent}pct": nested_display}),
                file_name=f"selected_nested_{pd.Timestamp.now().date()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.download_button(
                "📥 Download Selected (Excel)",
                data=save_df_to_excel_with_format(sel_display, sheet_name="Selected_Summary"),
                file_name=f"selected_products_{pd.Timestamp.now().date()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    st.markdown("----")
    st.info("Usage tips: Use Top% to find high-impact SKUs (Pareto). Use Bottom% to find long-tail / low-sales SKUs. Nested selection lets you drill into the top subset.")


# ------------------ COURIER BILL COUNT MODULE ------------------
elif st.session_state.page == "courier_mapper":

    import pandas as pd
    import streamlit as st
    from io import BytesIO
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    import re

    st.title("📦 Courier Bill Count Portal")

    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"

    # ============================================================
    # STEP 1: TRANSACTION FILE
    # ============================================================
    st.subheader("Step 1: Upload Transaction File")
    uploaded_file1 = st.file_uploader(
        "Upload first file (Transaction Details)",
        type=["csv", "xlsx"],
        key="file1"
    )

    df_grouped = None

    if uploaded_file1:
        df1 = pd.read_csv(uploaded_file1) if uploaded_file1.name.endswith(".csv") else pd.read_excel(uploaded_file1)
        df1.columns = df1.columns.str.strip()

        st.markdown("### 👀 Transaction File Preview")
        st.dataframe(df1.head(), width='stretch')

        required_cols1 = ["A/c No.", "Cust.Name", "Trn.No."]
        if not all(col in df1.columns for col in required_cols1):
            st.error(f"⚠️ First file must contain columns: {required_cols1}")
            st.stop()

        df_grouped = (
            df1.groupby(["A/c No.", "Cust.Name"], as_index=False)["Trn.No."]
            .count()
            .rename(columns={"Trn.No.": "No Of Bill"})
        )

    # ============================================================
    # STEP 2: COURIER FILE (OLD FORMAT LOGIC)
    # ============================================================
    st.subheader("Step 2: Upload Courier File")
    uploaded_file2 = st.file_uploader(
        "Upload second file (Courier Details)",
        type=["csv", "xlsx"],
        key="file2"
    )

    df2 = None

    if uploaded_file2 and uploaded_file1:
        df2 = pd.read_csv(uploaded_file2) if uploaded_file2.name.endswith(".csv") else pd.read_excel(uploaded_file2)

        df2.columns = (
            df2.columns
            .str.replace("\n", " ")
            .str.replace("\r", " ")
            .str.strip()
        )

        st.markdown("### 👀 Courier File Preview")
        st.dataframe(df2.head(), width='stretch')

        # Detect No Of Bill column (OLD LOGIC – UNCHANGED)
        bill_col_candidates = [col for col in df2.columns if "no of bill" in col.lower()]
        if len(bill_col_candidates) == 0:
            st.error("⚠️ Could not find 'No Of Bill' column in the second file!")
            st.stop()
        bill_col = bill_col_candidates[0]

        # Map No Of Bill (OLD LOGIC – UNCHANGED)
        df_grouped_unique = df_grouped.groupby("A/c No.", as_index=False)["No Of Bill"].sum()
        df2[bill_col] = df2["C.CODE"].map(
            df_grouped_unique.set_index("A/c No.")["No Of Bill"]
        )

    # ============================================================
    # STEP 3: TRAY FILE (NEW – ONLY ADDITION)
    # ============================================================
    st.subheader("Step 3: Upload Tray File")
    uploaded_file3 = st.file_uploader(
        "Upload tray file (Slip & Tray details)",
        type=["csv", "xlsx"],
        key="file3"
    )

    tray_map = {}

    if uploaded_file3:
        df3 = pd.read_csv(uploaded_file3) if uploaded_file3.name.endswith(".csv") else pd.read_excel(uploaded_file3)
        df3.columns = df3.columns.str.strip()

        st.markdown("### 👀 Tray File Preview")
        st.dataframe(df3.head(), width='stretch')

        if "CUSTNAME" not in df3.columns or "TRAYID" not in df3.columns:
            st.error("⚠️ Tray file must contain CUSTNAME and TRAYID columns")
            st.stop()

        # Extract customer code from [XXXX]
        df3["CUST_CODE"] = df3["CUSTNAME"].apply(
            lambda x: re.search(r"\[(\d+)\]", str(x)).group(1)
            if re.search(r"\[(\d+)\]", str(x)) else None
        )

        # Group tray IDs with pipe separator
        tray_map = (
            df3.dropna(subset=["CUST_CODE"])
            .groupby("CUST_CODE")["TRAYID"]
            .apply(lambda x: " | ".join(sorted(x.astype(str).unique())))
            .to_dict()
        )

    # ============================================================
    # FINAL OUTPUT (OLD FORMAT + TRAY ID COLUMN)
    # ============================================================
    if df2 is not None:

        # 🔹 ADD ONLY ONE COLUMN
        df2["TRAY ID"] = df2["C.CODE"].astype(str).map(tray_map).fillna("")

        st.success("✅ TRAY ID column added (format unchanged)")
        st.markdown("### ✅ Final Result Preview")
        st.dataframe(df2.head(10), width='stretch')

        # ------------------ Name & Date ------------------
        name_option = st.selectbox(
            "Select Name for report",
            ["BHUVANA", "KAVIYA", "KIRUBA", "SNEHA"]
        )
        current_date = datetime.now().strftime("%d.%m.%Y")

        # ------------------ Excel Creation (OLD FUNCTION – UNCHANGED) ------------------
        def create_excel_with_header(df, name, date_str):
            wb = Workbook()
            ws = wb.active

            # Title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            cell = ws.cell(row=1, column=1)
            cell.value = "YOGIRAM PHARMA COURIERS DETAILS"
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=14, bold=True)

            # Name & Date row
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
            ws.cell(row=2, column=1, value=f"NAME: {name}")
            ws.cell(row=2, column=len(df.columns), value=f"DATE {date_str}")

            # Header row
            for col_num, column_title in enumerate(df.columns, 1):
                ws.cell(row=3, column=col_num, value=column_title)
                ws.cell(row=3, column=col_num).font = Font(bold=True)
                ws.cell(row=3, column=col_num).alignment = Alignment(horizontal="center")

            # Data rows
            for row_num, row in enumerate(df.values, 4):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

            # Column widths (A4-friendly)
            for col_num, column_cells in enumerate(ws.columns, 1):
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

            return wb

        # ------------------ Download ------------------
        wb_final = create_excel_with_header(df2, name_option, current_date)
        output = BytesIO()
        wb_final.save(output)
        output.seek(0)

        st.download_button(
            "📥 Download Final Report (Excel)",
            data=output,
            file_name=f"Courier_Report_{current_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ------------------ PENDING ORDER LOCK ANALYZER ------------------
elif st.session_state.page == "pending_lock_analyzer":

    st.title("📦 Pending Order Lock Analyzer")

    if st.button("🏠 Back to Home"):
        go_home()

    st.subheader("Upload Required Files")

    pending_file = st.file_uploader(
        "Upload Pending Orders File",
        type=["xlsx","csv"],
        key="pending_file"
    )

    lock_file = st.file_uploader(
        "Upload Customer Lock File",
        type=["xlsx","csv"],
        key="lock_file"
    )

    salesman_file = st.file_uploader(
        "Upload Salesman Mapping File",
        type=["xlsx","csv"],
        key="salesman_file"
    )

    if pending_file and lock_file and salesman_file:

        # ---------- READ FILES ----------
        pending_df = pd.read_excel(pending_file) if pending_file.name.endswith(".xlsx") else pd.read_csv(pending_file)
        lock_df = pd.read_excel(lock_file) if lock_file.name.endswith(".xlsx") else pd.read_csv(lock_file)
        sales_df = pd.read_excel(salesman_file) if salesman_file.name.endswith(".xlsx") else pd.read_csv(salesman_file)

        # ---------- CLEAN COLUMN NAMES ----------
        pending_df.columns = pending_df.columns.str.strip()
        lock_df.columns = lock_df.columns.str.strip()
        sales_df.columns = sales_df.columns.str.strip()

        # ---------- VALIDATE REQUIRED COLUMNS ----------
        if "Ord No." not in pending_df.columns or "Code" not in pending_df.columns:
            st.error("Pending file must contain 'Ord No.' and 'Code'")
            st.stop()

        if "Code" not in lock_df.columns or "Reason" not in lock_df.columns:
            st.error("Customer Lock file must contain 'Code' and 'Reason'")
            st.stop()

        if "CODE" not in sales_df.columns or "SALES MAN" not in sales_df.columns:
            st.error("Salesman file must contain 'CODE' and 'SALES MAN'")
            st.stop()

        # ---------- FIX ORD NO FORMAT ----------
        pending_df["Ord No."] = pd.to_numeric(pending_df["Ord No."], errors="coerce")

        # ---------- REMOVE DUPLICATE CODES ----------
        lock_df = lock_df.drop_duplicates(subset="Code", keep="first")
        sales_df = sales_df.drop_duplicates(subset="CODE", keep="first")

        # ---------- LOOKUP REASON ----------
        pending_df["Reason"] = pending_df["Code"].map(
            lock_df.set_index("Code")["Reason"]
        )

        # ---------- LOOKUP SALESMAN ----------
        pending_df["Salesman"] = pending_df["Code"].map(
            sales_df.set_index("CODE")["SALES MAN"]
        )

        # ---------- REPORT 1 (ALL DATA) ----------
        report_all = pending_df.copy()

        # ---------- REPORT 2 (REMOVE 0 ORDERS) ----------
        report_without_zero = pending_df[pending_df["Ord No."] != 0].copy()

        # ---------- PREVIEW ----------
        st.subheader("📄 Report 1 — All Orders (Includes 0)")
        st.dataframe(report_all, width='stretch')

        st.subheader("📄 Report 2 — Orders Without 0")
        st.dataframe(report_without_zero, width='stretch')

        # ---------- EXCEL FUNCTION ----------
        def to_excel(df):
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
            return output.getvalue()

        # ---------- DOWNLOAD BUTTONS ----------
        st.markdown("---")

        st.download_button(
            "📥 Download Report 1 (All Orders)",
            data=to_excel(report_all),
            file_name=f"Pending_Orders_All_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            "📥 Download Report 2 (Without 0 Orders)",
            data=to_excel(report_without_zero),
            file_name=f"Pending_Orders_No_Zero_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
# --------------------------- APOLLO CHECK ---------------------------
elif st.session_state.page == "apollo":
    st.title("🚀 Apollo Check Module")
    if st.button("🏠 Back to Home"): go_home()

    # ------------------ FILE UPLOADER ------------------

    if "apollo_file" not in st.session_state:
        st.session_state.apollo_file = None

    uploaded_file = st.file_uploader("Upload your Excel/CSV file for Apollo Check", type=["xlsx", "csv"])
    if uploaded_file:
        st.session_state.apollo_file = uploaded_file

    if not st.session_state.apollo_file:
        st.info("Please upload a file to continue.")
        st.stop()

    # ------------------ READ FILE ------------------
    df = pd.read_excel(st.session_state.apollo_file) if st.session_state.apollo_file.name.endswith(".xlsx") else pd.read_csv(st.session_state.apollo_file)
    df['INDENT DATE'] = pd.to_datetime(df['INDENT DATE'], errors='coerce')
    df = df.sort_values(['SHOP NAME','ITEM NAME','INDENT DATE'])

    # ------------------ TABLE 1: CONTINUOUS ORDERS ------------------
    st.subheader("📊 Continuous Orders Table")
    consecutive_days = st.selectbox("Select consecutive days", list(range(1,16)), key="consec_days")
    continuous_list = []

    for shop, shop_df in df.groupby('SHOP NAME'):
        for item, item_df in shop_df.groupby('ITEM NAME'):
            dates = item_df['INDENT DATE'].sort_values().tolist()
            count = 1
            for i in range(1,len(dates)):
                if (dates[i] - dates[i-1]).days == 1:
                    count += 1
                else:
                    count = 1
                if count >= consecutive_days:
                    continuous_list.append([shop, item, dates[i-consecutive_days+1], dates[i], consecutive_days])
                    break

    continuous_df = pd.DataFrame(continuous_list, columns=['Shop Name','Item Name','Start Date','End Date','Consecutive Days'])
    st.dataframe(continuous_df)

    # ------------------ TABLE 2: NON-CONTINUOUS ORDERS ------------------
    st.subheader("📊 Non-Continuous Orders Table")
    total_days = st.selectbox("Select total order days", list(range(1,16)), key="total_days")
    non_continuous_list = []

    for shop, shop_df in df.groupby('SHOP NAME'):
        for item, item_df in shop_df.groupby('ITEM NAME'):
            unique_days = item_df['INDENT DATE'].nunique()
            if unique_days >= total_days:
                non_continuous_list.append([shop, item, unique_days])

    non_continuous_df = pd.DataFrame(non_continuous_list, columns=['Shop Name','Item Name','Total Days Ordered'])
    st.dataframe(non_continuous_df)

    # ------------------ COMPLETE SUMMARY TABLE ------------------
    st.subheader("📋 Complete Summary Table (All Data)")
    summary_list = []

    for shop, shop_df in df.groupby('SHOP NAME'):
        for item, item_df in shop_df.groupby('ITEM NAME'):
            summary_list.append([shop, item, item_df['INDENT DATE'].nunique()])

    summary_df = pd.DataFrame(summary_list, columns=['Shop Name','Item Name','Total Days Ordered'])
    summary_df = summary_df.sort_values('Total Days Ordered', ascending=False)
    st.dataframe(summary_df)


# ------------------ DUE LIST CHECKER MODULE ------------------
elif st.session_state.page == "aging_analysis":
    import pandas as pd
    import streamlit as st
    from io import BytesIO
    from datetime import datetime

    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"

    st.title("⏳ Due List Checker Portal")



    # File Uploader
    uploaded_file = st.file_uploader("Upload your Dues Statement (Excel/CSV)", type=["xlsx", "csv"], key="aging_upload")

    if uploaded_file:
        try:
            # Using the new progress bar helper
            df_raw = load_data_with_progress(uploaded_file)
            
            st.success("✅ File processed successfully!")
            
            # Column Normalization
            df_raw.columns = df_raw.columns.str.strip()
            
            # Check for required columns
            required_cols = ["REL NO.", "PARTYNAME", "SALESMAN", "DAYS", "OSAMT"]
            missing = [c for c in required_cols if c not in df_raw.columns]
            
            if missing:
                st.error(f"⚠️ Missing columns: {', '.join(missing)}")
                st.stop()
            
            # Data Cleaning & Type Casting (Crucial to prevent Arrow error)
            df_raw["REL NO."] = df_raw["REL NO."].astype(str).str.strip()
            df_raw["PARTYNAME"] = df_raw["PARTYNAME"].astype(str).str.strip()
            df_raw["SALESMAN"] = df_raw["SALESMAN"].astype(str).str.strip()
            df_raw["OSAMT"] = pd.to_numeric(df_raw["OSAMT"], errors="coerce").fillna(0)
            df_raw["DAYS"] = pd.to_numeric(df_raw["DAYS"], errors="coerce").fillna(0)

            # --- CUSTOM BUCKETING LOGIC ---
            def categorize_days(d):
                if d <= 30: return "0-30 Days"
                elif d <= 60: return "30-60 Days"
                elif d <= 90: return "60-90 Days"
                elif d <= 120: return "90-120 Days"
                elif d <= 180: return "120-180 Days"
                elif d <= 360: return "180-360 Day"
                else: return "Above 360 Day"

            df_raw["Due Bucket"] = df_raw["DAYS"].apply(categorize_days)
            
            # Create Pivot Table
            pivot_df = pd.pivot_table(
                df_raw,
                index=["REL NO.", "PARTYNAME", "SALESMAN"],
                columns="Due Bucket",
                values="OSAMT",
                aggfunc="sum",
                fill_value=0
            )
            
            # Order Buckets
            all_buckets = [
                "0-30 Days", "30-60 Days", "60-90 Days", 
                "90-120 Days", "120-180 Days", "180-360 Day", "Above 360 Day"
            ]
            
            # Ensure all buckets exist
            for b in all_buckets:
                if b not in pivot_df.columns:
                    pivot_df[b] = 0
            
            pivot_df = pivot_df[all_buckets]
            
            # Add Grand Total column
            pivot_df["Grand Total"] = pivot_df.sum(axis=1)
            
            # --- ADD TOTAL ROW AT THE BOTTOM ---
            # Calculate column totals
            col_totals = pivot_df.sum(numeric_only=True).to_frame().T
            # Create a MultiIndex for the total row to match the pivot_df
            col_totals.index = pd.MultiIndex.from_tuples([("", "--- GRAND TOTAL ---", "")], names=["REL NO.", "PARTYNAME", "SALESMAN"])
            
            # Combine
            final_display_df = pd.concat([pivot_df, col_totals])
            
            st.write("### 📊 Complete Aging Pivot Table")
            
            # Apply Premium Styling
            def style_aging_table(styler):
                styler.format("₹{:,.2f}")
                # Add background gradient to buckets (excluding the Grand Total label row if possible, but simpler to apply to all)
                styler.background_gradient(cmap="YlOrRd", subset=all_buckets)
                styler.set_properties(**{'background-color': '#161b22', 'color': '#ffffff', 'border-color': '#30363d'})
                return styler

            st.dataframe(final_display_df.style.pipe(style_aging_table), width='stretch')
            
            # Full Download Option (We download the pivot_df without the styled TOTAL row for clean data usage)
            full_excel = save_df_to_excel_with_format(final_display_df, sheet_name="Full Aging Report")
            st.download_button(
                label="📥 Download Full Result (Excel)",
                data=full_excel,
                file_name=f"Aging_Analysis_Full_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.markdown("---")
            
            # Salesman-wise Filtered Download
            st.subheader("🎯 Salesman-wise Selective Download")
            
            # Get unique salesmen
            salesmen_list = sorted(df_raw["SALESMAN"].dropna().unique())
            
            # Use a single bar for selection
            # We add a "--- ALL SALESMEN ---" option at the top for quick bulk selection
            options_with_all = ["--- SELECT ALL SALESMEN ---"] + salesmen_list

            selected_raw = st.multiselect(
                "🔍 Type to search and select salesmen (Press Enter to add)",
                options=options_with_all,
                default=st.session_state.get("selected_salesmen", []),
                placeholder="Search names (e.g. 'Gokul')..."
            )
            
            # Logic to handle "Select All"
            if "--- SELECT ALL SALESMEN ---" in selected_raw:
                selected_salesmen = salesmen_list
            else:
                selected_salesmen = selected_raw
            
            # Sync session state
            st.session_state.selected_salesmen = selected_salesmen
            
            if selected_salesmen:
                # Filter the pivot table (using the original pivot_df before the total row was added)
                filtered_pivot = pivot_df.reset_index()
                filtered_pivot = filtered_pivot[filtered_pivot["SALESMAN"].isin(selected_salesmen)]
                
                # Set index back
                filtered_pivot = filtered_pivot.set_index(["REL NO.", "PARTYNAME", "SALESMAN"])
                
                # Add Total Row for this filtered set
                f_col_totals = filtered_pivot.sum(numeric_only=True).to_frame().T
                f_col_totals.index = pd.MultiIndex.from_tuples([("", "--- FILTERED TOTAL ---", "")], names=["REL NO.", "PARTYNAME", "SALESMAN"])
                filtered_display_df = pd.concat([filtered_pivot, f_col_totals])
                
                st.write(f"### 📋 Preview for Selected Salesmen ({len(selected_salesmen)})")
                st.dataframe(filtered_display_df.style.pipe(style_aging_table), width='stretch')
                
                # File Naming Logic
                salesmen_str = "_".join(selected_salesmen)[:50] 
                today_str = datetime.now().strftime("%Y-%m-%d")
                filtered_filename = f"dues list for ({salesmen_str}) {today_str}.xlsx"
                
                filtered_excel = save_df_to_excel_with_format(filtered_display_df, sheet_name="Selective Dues")
                st.download_button(
                    label=f"📥 Download Dues List for Selected",
                    data=filtered_excel,
                    file_name=filtered_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("Select one or more salesmen above to download their specific dues list.")

        except Exception as e:
            st.error(f"⚠️ Error processing file: {e}")



# ------------------ PHARMA FORECAST MODULE ------------------
elif st.session_state.page == "pharma_forecast":
    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"
    
    # Custom CSS for Forecast components to match Aurora theme
    st.markdown("""
        <style>
        .forecast-hero {
            background: rgba(255,255,255,0.03) !important;
            border: 1px solid rgba(0,242,255,0.2) !important;
            border-radius: 20px !important;
            padding: 2rem !important;
            margin-bottom: 2rem !important;
            backdrop-filter: blur(10px);
        }
        .forecast-title {
            font-size: 2.2rem !important;
            font-weight: 800 !important;
            background: linear-gradient(135deg, #fff, #00f2ff) !important;
            -webkit-background-clip: text !important;
            -webkit-text-fill-color: transparent !important;
        }
        .forecast-subtitle {
            color: rgba(255,255,255,0.5) !important;
            font-size: 1rem !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    if render_forecast_module:
        render_forecast_module()
    else:
        st.error("Forecast Module not found in pharma_forecast_app folder.")

# ------------------ ADMIN LOGIN LOG PAGE ------------------
elif st.session_state.page == "admin_log":
    st.title("📝 Login Activity Log")
    if st.button("🏠 Back to Home"): go_home()

    if os.path.exists(LOGIN_LOG_FILE):
        log_df = pd.read_csv(LOGIN_LOG_FILE)
        st.dataframe(log_df)

        def to_excel(df):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='LoginLogs')
            return output.getvalue()

        st.download_button("📥 Download Login Logs", data=to_excel(log_df), file_name="LoginLogs.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("No login records yet.")



# END OF FILE   
