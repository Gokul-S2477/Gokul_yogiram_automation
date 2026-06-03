import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


# ─────────────────────────────────────────────
# Helper: read any uploaded file → DataFrame
# ─────────────────────────────────────────────
def _read_file(f):
    if f is None:
        return None
    if f.name.lower().endswith(".csv"):
        return pd.read_csv(f)
    else:
        return pd.read_excel(f)


# ─────────────────────────────────────────────
# Helper: standardize column names
# ─────────────────────────────────────────────
def _std_cols(df):
    df.columns = df.columns.str.strip().str.upper()
    return df


# ─────────────────────────────────────────────
# Core: map gold code + filter bad items
# ─────────────────────────────────────────────
def _process_branch(df_main, df_master, main_code_col, master_code_col):
    """
    1. Map gold code from master onto main via code columns.
    2. Remove rows where Item name starts with # or ~.
    Returns cleaned df with gold code prepended.
    """
    df_main = df_main.copy()
    df_master = df_master.copy()

    # Standardize
    df_main = _std_cols(df_main)
    df_master = _std_cols(df_master)

    main_code_col   = main_code_col.strip().upper()
    master_code_col = master_code_col.strip().upper()

    # Verify columns exist
    if main_code_col not in df_main.columns:
        raise ValueError(
            f"Column '{main_code_col}' not found in main file. "
            f"Available: {list(df_main.columns)}"
        )

    gc_candidates = [c for c in df_master.columns if "GOLD" in c]
    if not gc_candidates:
        raise ValueError(
            "No 'GOLD CODE' column found in item master. "
            f"Available: {list(df_master.columns)}"
        )
    gold_col_in_master = gc_candidates[0]

    if master_code_col not in df_master.columns:
        raise ValueError(
            f"Column '{master_code_col}' not found in item master. "
            f"Available: {list(df_master.columns)}"
        )

    # Cast join keys to str and strip whitespace for safety before lookup/deduplication
    df_main[main_code_col] = df_main[main_code_col].astype(str).str.strip()
    df_master[master_code_col] = df_master[master_code_col].astype(str).str.strip()

    # Build lookup: master_code → gold_code (deduplicate to avoid InvalidIndexError)
    df_master_dedup = df_master.drop_duplicates(subset=[master_code_col], keep='first')
    lookup = df_master_dedup.set_index(master_code_col)[gold_col_in_master].astype(str)

    df_main["GOLD CODE"] = df_main[main_code_col].map(lookup)

    # Remove items starting with # or ~
    item_col = next(
        (c for c in df_main.columns if c in ["ITEM", "ITEM NAME", "ITEMNAME"]),
        None
    )
    if item_col:
        mask = df_main[item_col].astype(str).str.strip().str.startswith(("#", "~"))
        removed = mask.sum()
        df_main = df_main[~mask].copy()
        if removed:
            st.info(f"ℹ️ Removed **{removed}** items starting with # or ~ from this branch.")

    return df_main


# ─────────────────────────────────────────────
# Core: pivot on GOLD CODE
# ─────────────────────────────────────────────
MONTH_COLS = ["APR", "MAY", "JUN", "JUL", "AUG", "SEP",
              "OCT", "NOV", "DEC", "JAN", "FEB", "MAR"]

VALUE_COLS_PRIORITY = [
    "NET SALE", "SALE", "TRR", "SS", "SP", "PUR.", "PUR RET",
    "SPUR RET", "TRI", "SRET", "ADJ.", "CLS.STK", "OP.", "CLS STK"
]

def _build_pivot(combined_df, value_col):
    """
    Pivot: strictly uses the requested columns:
    gold code, Company, Item, Pack, Mar, Adj., Sale, SP, SPur Ret, Pur Ret, SS, Apr, Op., TRR, Pur., TRI, SRet, Cls.Stk
    """
    combined_df = combined_df.copy()
    combined_df.columns = combined_df.columns.str.strip().str.upper()

    def find_col(df_cols, candidates):
        for c in candidates:
            if c in df_cols:
                return c
        return None

    # Mappings from requested names to potential column headers in the files
    col_mappings = {
        "gold code": ["GOLD CODE"],
        "Company": ["COMPANY", "COMPCODE", "COMP"],
        "Item": ["ITEM", "ITEM NAME", "ITEMNAME"],
        "Pack": ["PACK", "PPACK", "PACKSIZE", "PACK SIZE"],
        "Mar": ["MAR", "MARCH"],
        "Adj.": ["ADJ.", "ADJ", "ADJUSTMENT"],
        "Sale": ["SALE", "NET SALE", "NET SALES", "SALES"],
        "SP": ["SP"],
        "SPur Ret": ["SPUR RET", "SPUR.RET", "SPUR. RET", "SPUR RET."],
        "Pur Ret": ["PUR RET", "PUR.RET", "PUR. RET", "PUR RET.", "PURCHASE RETURN", "PURCHASE RETURNS"],
        "SS": ["SS"],
        "Apr": ["APR", "APRIL"],
        "Op.": ["OP.", "OP", "OPENING", "OPENING STOCK", "OP.STK", "OP STK"],
        "TRR": ["TRR"],
        "Pur.": ["PUR.", "PUR", "PURCHASE", "PURCHASES"],
        "TRI": ["TRI"],
        "SRet": ["SRET", "SRET.", "SALE RETURN", "SALES RETURN", "SALES RETURNS"],
        "Cls.Stk": ["CLS.STK", "CLS STK", "CLS. STK", "CLOSING", "CLOSING STOCK", "CLS STK."]
    }

    actual_cols = {}
    for target_name, candidates in col_mappings.items():
        match = find_col(combined_df.columns, candidates)
        if match:
            actual_cols[target_name] = match

    # Ensure reference columns exist (create if missing)
    ref_targets = ["gold code", "Company", "Item", "Pack"]
    for target in ref_targets:
        actual_name = actual_cols.get(target)
        if not actual_name:
            col_upper = target.upper()
            combined_df[col_upper] = ""
            actual_cols[target] = col_upper

    # Ensure value columns exist (create with 0.0 if missing, else convert to numeric)
    value_targets = ["Mar", "Adj.", "Sale", "SP", "SPur Ret", "Pur Ret", "SS", "Apr", "Op.", "TRR", "Pur.", "TRI", "SRet", "Cls.Stk"]
    for target in value_targets:
        actual_name = actual_cols.get(target)
        if actual_name:
            combined_df[actual_name] = pd.to_numeric(combined_df[actual_name], errors="coerce").fillna(0)
        else:
            col_upper = target.upper()
            combined_df[col_upper] = 0.0
            actual_cols[target] = col_upper

    # Extract unique reference info prioritizing Branch 1
    actual_gold_code = actual_cols["gold code"]
    actual_ref_cols = [actual_cols["Company"], actual_cols["Item"], actual_cols["Pack"]]

    if "_BRANCH" in combined_df.columns:
        b1_mask = combined_df["_BRANCH"].astype(str).str.strip().str.upper() == "BRANCH 1"
        b1_rows = combined_df[b1_mask]
        b2_rows = combined_df[~b1_mask]
    else:
        b1_rows = combined_df
        b2_rows = pd.DataFrame(columns=combined_df.columns)

    ref_b1 = b1_rows[[actual_gold_code] + actual_ref_cols].drop_duplicates(subset=[actual_gold_code])
    ref_b2 = b2_rows[[actual_gold_code] + actual_ref_cols].drop_duplicates(subset=[actual_gold_code])
    
    # Union reference details, keeping Branch 1 first
    ref_df = pd.concat([ref_b1, ref_b2], ignore_index=True).drop_duplicates(subset=[actual_gold_code], keep="first")

    # Group strictly by GOLD CODE and sum values
    actual_val_cols = [actual_cols[v] for v in value_targets]
    pivot_vals = combined_df.groupby(actual_gold_code, as_index=False)[actual_val_cols].sum()

    # Merge reference columns back onto the summed monthly values
    pivot = pd.merge(ref_df, pivot_vals, on=actual_gold_code, how="right")

    # Rename columns to their exact requested casing
    rename_dict = {actual: target for target, actual in actual_cols.items()}
    pivot = pivot.rename(columns=rename_dict)

    # Order columns exactly as requested by the user
    target_order = ["gold code", "Company", "Item", "Pack"] + value_targets
    pivot = pivot[target_order]

    return pivot, value_targets


# ─────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────
def _export_to_excel(combined_df, pivot_df, month_cols):
    wb = Workbook()

    orange_fill  = PatternFill("solid", fgColor="F4B084")
    blue_fill    = PatternFill("solid", fgColor="BDD7EE")
    green_fill   = PatternFill("solid", fgColor="C6EFCE")
    bold_font    = Font(bold=True, name="Calibri", size=10)
    header_font  = Font(bold=True, name="Calibri", size=10, color="000000")
    data_font    = Font(name="Calibri", size=10)
    thin         = Side(style="thin", color="000000")
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")
    right_align  = Alignment(horizontal="right",  vertical="center")

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    def write_df_to_sheet(ws, df, fill, title=None):
        start_row = 1
        if title:
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=len(df.columns)
            )
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.fill = PatternFill("solid", fgColor="1F3864")
            title_cell.font = Font(bold=True, name="Calibri", size=12, color="FFFFFF")
            title_cell.alignment = center_align
            ws.row_dimensions[1].height = 20
            start_row = 2

        # Header
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=ci, value=col)
            cell.fill = fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Data
        for ri, row_vals in enumerate(df.itertuples(index=False), start_row + 1):
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = data_font
                cell.border = thin_border
                col_name = str(df.columns[ci - 1]).upper()
                if col_name in ["GOLD CODE", "CODE", "COMPCODE"]:
                    cell.alignment = center_align
                elif isinstance(val, (int, float, np.integer, np.floating)):
                    cell.alignment = right_align
                    cell.number_format = "#,##0.00"
                else:
                    cell.alignment = left_align
        auto_width(ws)

    # Sheet 1: Pivot
    ws1 = wb.active
    ws1.title = "Pivot by Gold Code"
    write_df_to_sheet(ws1, pivot_df, green_fill, title="Sun Statement – Gold Code Pivot")

    # Sheet 2: Combined Raw
    export_combined = combined_df.copy()
    if "_BRANCH" in export_combined.columns:
        export_combined = export_combined.drop(columns=["_BRANCH"])

    ws2 = wb.create_sheet("Combined Data")
    write_df_to_sheet(ws2, export_combined, orange_fill, title="Combined Data")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# Main render function
# ─────────────────────────────────────────────
def render_sun_statement():
    st.title("☀️ Sun Statement Module")

    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"
        st.rerun()

    st.markdown("""
    <p style='color:rgba(255,255,255,0.55); font-size:0.95rem; margin-bottom:1.5rem;'>
        Upload the <strong>Sun Statement</strong> and <strong>Item Master</strong> files for two branches.
        The system will map <em>Gold Codes</em>, remove invalid items, merge both branches,
        and generate a <strong>Gold Code Pivot</strong> ready for download.
    </p>
    """, unsafe_allow_html=True)

    # ── Tabs ──
    tab_upload, tab_result = st.tabs(["📂 Upload Files", "📊 Results & Download"])

    with tab_upload:
        st.markdown("### 🌿 Branch 1")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**📄 Branch 1 – Sun Statement (Main File)**")
            b1_main = st.file_uploader(
                "Upload Branch 1 Main File",
                type=["xlsx", "xls", "csv"],
                key="b1_main",
                help="The SUN Statement export with columns: Code, Item, Company, months…"
            )
        with col2:
            st.markdown("**📋 Branch 1 – Item Master (with Gold Code)**")
            b1_master = st.file_uploader(
                "Upload Branch 1 Item Master",
                type=["xlsx", "xls", "csv"],
                key="b1_master",
                help="Must contain a 'Gold Code' column and the item code column."
            )

        b1_main_code   = st.text_input("Branch 1 – Code column in Main File",   value="CODE",     key="b1mc")
        b1_master_code = st.text_input("Branch 1 – Code column in Item Master", value="PPPLCODE", key="b1sc")

        st.divider()

        st.markdown("### 🌾 Branch 2")
        col3, col4 = st.columns(2)
        with col3:
            st.markdown("**📄 Branch 2 – Sun Statement (Main File)**")
            b2_main = st.file_uploader(
                "Upload Branch 2 Main File",
                type=["xlsx", "xls", "csv"],
                key="b2_main"
            )
        with col4:
            st.markdown("**📋 Branch 2 – Item Master (with Gold Code)**")
            b2_master = st.file_uploader(
                "Upload Branch 2 Item Master",
                type=["xlsx", "xls", "csv"],
                key="b2_master"
            )

        b2_main_code   = st.text_input("Branch 2 – Code column in Main File",   value="CODE",     key="b2mc")
        b2_master_code = st.text_input("Branch 2 – Code column in Item Master", value="PPPLCODE", key="b2sc")

        st.divider()

        st.markdown("### ⚙️ Pivot Settings")
        value_col_input = st.text_input(
            "Value column for pivot (leave blank to auto-detect)",
            value="",
            key="val_col",
            help="e.g. NET SALE, SALE, TRR … If blank, first numeric month column is used."
        )

        process_btn = st.button("🚀 Process & Generate Pivot", use_container_width=True, key="process_btn")

    with tab_result:
        if "sun_stmt_result" not in st.session_state:
            st.info("Upload files in the **Upload Files** tab and click **Process** to see results here.")

    # ── Process ──
    if process_btn:
        # Validation
        errors = []
        if not b1_main:   errors.append("Branch 1 Main File is missing.")
        if not b1_master: errors.append("Branch 1 Item Master is missing.")
        if not b2_main:   errors.append("Branch 2 Main File is missing.")
        if not b2_master: errors.append("Branch 2 Item Master is missing.")

        if errors:
            for e in errors:
                st.error(f"❌ {e}")
        else:
            with st.spinner("🔄 Processing Sun Statement..."):
                try:
                    # Read all 4 files
                    df_b1_main   = _read_file(b1_main)
                    df_b1_master = _read_file(b1_master)
                    df_b2_main   = _read_file(b2_main)
                    df_b2_master = _read_file(b2_master)

                    st.markdown("---")
                    st.markdown("#### Branch 1 Processing")
                    df_b1 = _process_branch(df_b1_main, df_b1_master, b1_main_code, b1_master_code)
                    df_b1["_BRANCH"] = "Branch 1"

                    st.markdown("#### Branch 2 Processing")
                    df_b2 = _process_branch(df_b2_main, df_b2_master, b2_main_code, b2_master_code)
                    df_b2["_BRANCH"] = "Branch 2"

                    # Combine: Branch 1 on top, Branch 2 below
                    combined = pd.concat([df_b1, df_b2], ignore_index=True)

                    # Build pivot
                    pivot_df, found_months = _build_pivot(combined, value_col_input)

                    # Cache in session
                    st.session_state["sun_stmt_result"] = {
                        "combined": combined,
                        "pivot":    pivot_df,
                        "months":   found_months,
                    }

                    st.success("✅ Processing complete! View results in the **Results & Download** tab.")

                except Exception as ex:
                    st.error(f"⚠️ Error during processing: {ex}")
                    st.exception(ex)

    # ── Results tab rendering ──
    with tab_result:
        if "sun_stmt_result" in st.session_state:
            res = st.session_state["sun_stmt_result"]
            pivot_df = res["pivot"]
            combined = res["combined"]
            months   = res["months"]

            # ── Pivot Preview ──
            st.subheader("📊 Gold Code Pivot")
            st.caption(f"Columns: {', '.join(pivot_df.columns)}")

            # Style pivot: highlight TOTAL column
            def highlight_total(df):
                styles = pd.DataFrame("", index=df.index, columns=df.columns)
                if "TOTAL" in df.columns:
                    styles["TOTAL"] = "background-color: #1e3a5f; color: #7ecfff; font-weight: bold;"
                return styles

            st.dataframe(
                pivot_df.style.apply(highlight_total, axis=None).format(
                    {c: "{:,.2f}" for c in months + (["TOTAL"] if "TOTAL" in pivot_df.columns else [])}
                ),
                use_container_width=True,
                height=480
            )

            st.markdown("---")

            # ── Combined Raw Preview ──
            with st.expander("🔍 View Combined Raw Data", expanded=False):
                display_combined = combined.copy()
                if "_BRANCH" in display_combined.columns:
                    display_combined = display_combined.drop(columns=["_BRANCH"])
                st.dataframe(display_combined, use_container_width=True, height=350)

            st.markdown("---")

            # ── Gold Code unmapped summary ──
            unmapped = pivot_df[pivot_df["gold code"].astype(str).isin(["nan", "None", "", "NaN"])]
            if not unmapped.empty:
                st.warning(
                    f"⚠️ **{len(unmapped)} rows** could not be matched to a Gold Code. "
                    "Check that the code columns match between files."
                )

            # ── Download ──
            st.subheader("📥 Download Results")
            excel_bytes = _export_to_excel(combined, pivot_df, months)

            st.download_button(
                label="📥 Download Sun Statement Excel",
                data=excel_bytes,
                file_name="sun_statement_pivot.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # ── Per-company breakdown ──
            if "Company" in pivot_df.columns:
                st.markdown("---")
                st.subheader("🏢 Company-wise Totals")
                comp_col = "Company"
                comp_pivot = pivot_df.groupby(comp_col)[months].sum()
                comp_pivot["TOTAL"] = comp_pivot.sum(axis=1)
                comp_pivot = comp_pivot.sort_values("TOTAL", ascending=False)
                st.dataframe(
                    comp_pivot.style.format("{:,.2f}").background_gradient(
                        cmap="Blues", subset=["TOTAL"]
                    ),
                    use_container_width=True,
                )
