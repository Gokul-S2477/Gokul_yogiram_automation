import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def render_branch_transfer():
    st.title("🔄 Branch Transfer (BT) Module")

    if st.button("🏠 Back to Home"):
        st.session_state.page = "home"
        st.rerun()

    st.markdown("""
        Upload the required files for the Main Branch and other branches. 
        The system will calculate Required Stock based on Sales files and identify Excess/Deficit.
        Then, it will recommend stock transfers from branches with excess stock to the Main Branch.
    """)

    target_days = st.number_input("Target Inventory Days", min_value=1, value=30, step=1)
    
    num_branches = st.number_input("Number of Branches to Map", min_value=1, max_value=20, value=2, step=1)
    
    st.info("Branch 1 is always the **Main Branch**. It requires stock. Other branches provide stock.")
    
    # Store branch data
    branch_data = []

    with st.form("bt_upload_form"):
        for i in range(int(num_branches)):
            st.markdown(f"### Branch {i+1} {'(Main Branch)' if i==0 else ''}")
            branch_name = st.text_input(f"Branch Name", key=f"b_name_{i}", value="Main Branch" if i==0 else f"Branch {i+1}")
            
            c1, c2, c3 = st.columns(3)
            with c1:
                stock_file = st.file_uploader(f"Current Stock File", key=f"stock_{i}", type=["csv", "xlsx", "xls"])
            with c2:
                sales1_file = st.file_uploader(f"Sales File 1", key=f"s1_{i}", type=["csv", "xlsx", "xls"])
                sales1_days = st.number_input(f"Days for Sales 1", key=f"s1d_{i}", min_value=1, value=90)
            with c3:
                sales2_file = st.file_uploader(f"Sales File 2", key=f"s2_{i}", type=["csv", "xlsx", "xls"])
                sales2_days = st.number_input(f"Days for Sales 2", key=f"s2d_{i}", min_value=1, value=45)
            
            st.divider()
            
            branch_data.append({
                "name": branch_name,
                "stock_file": stock_file,
                "s1_file": sales1_file,
                "s1_days": sales1_days,
                "s2_file": sales2_file,
                "s2_days": sales2_days,
                "is_main": (i == 0)
            })
            
        submitted = st.form_submit_button("Process & Generate Recommendations", use_container_width=True)

    if submitted:
        # Validate files
        valid = True
        for b in branch_data:
            if not b["stock_file"] or not b["s1_file"] or not b["s2_file"]:
                st.error(f"Missing files for {b['name']}. Please upload all 3 files.")
                valid = False
                break
        
        if valid:
            with st.spinner("Processing Branch Data..."):
                processed_branches = {}
                
                for b in branch_data:
                    # Read files
                    try:
                        if b["stock_file"].name.endswith(".csv"):
                            df_stock = pd.read_csv(b["stock_file"])
                        else:
                            df_stock = pd.read_excel(b["stock_file"])
                            
                        if b["s1_file"].name.endswith(".csv"):
                            df_s1 = pd.read_csv(b["s1_file"])
                        else:
                            df_s1 = pd.read_excel(b["s1_file"])
                            
                        if b["s2_file"].name.endswith(".csv"):
                            df_s2 = pd.read_csv(b["s2_file"])
                        else:
                            df_s2 = pd.read_excel(b["s2_file"])
                    except Exception as e:
                        st.error(f"Error reading files for {b['name']}: {e}")
                        return
                    
                    # Standardize columns (strip spaces, uppercase)
                    df_stock.columns = df_stock.columns.str.strip().str.upper()
                    df_s1.columns = df_s1.columns.str.strip().str.upper()
                    df_s2.columns = df_s2.columns.str.strip().str.upper()
                    
                    stock_item_col = "ITEM CODE" if "ITEM CODE" in df_stock.columns else "BARCODE"
                    gold_code_col = "GOLD CODE" if "GOLD CODE" in df_stock.columns else "GOLDCODE"
                    stock_qty_col = "QTY" if "QTY" in df_stock.columns else "QTY."
                    stock_name_col = "ITEM NAME" if "ITEM NAME" in df_stock.columns else "NAME"
                    
                    s1_item_col = "BARCODE" if "BARCODE" in df_s1.columns else "ITEM CODE"
                    s1_qty_col = "QTY." if "QTY." in df_s1.columns else "QTY"
                    
                    s2_item_col = "BARCODE" if "BARCODE" in df_s2.columns else "ITEM CODE"
                    s2_qty_col = "QTY." if "QTY." in df_s2.columns else "QTY"
                    
                    # Drop duplicates in sales just in case by summing up
                    df_s1 = df_s1.groupby(s1_item_col)[s1_qty_col].sum().reset_index()
                    df_s1.rename(columns={s1_qty_col: "SALES_1_QTY", s1_item_col: "S1_ITEM_CODE"}, inplace=True)
                    
                    df_s2 = df_s2.groupby(s2_item_col)[s2_qty_col].sum().reset_index()
                    df_s2.rename(columns={s2_qty_col: "SALES_2_QTY", s2_item_col: "S2_ITEM_CODE"}, inplace=True)
                    
                    # Merge sales into stock safely
                    df_stock = pd.merge(df_stock, df_s1, left_on=stock_item_col, right_on="S1_ITEM_CODE", how="left")
                    if "S1_ITEM_CODE" in df_stock.columns:
                        df_stock.drop(columns=["S1_ITEM_CODE"], inplace=True)
                        
                    df_stock = pd.merge(df_stock, df_s2, left_on=stock_item_col, right_on="S2_ITEM_CODE", how="left")
                    if "S2_ITEM_CODE" in df_stock.columns:
                        df_stock.drop(columns=["S2_ITEM_CODE"], inplace=True)
                        
                    df_stock["SALES_1_QTY"] = pd.to_numeric(df_stock["SALES_1_QTY"], errors='coerce').fillna(0)
                    df_stock["SALES_2_QTY"] = pd.to_numeric(df_stock["SALES_2_QTY"], errors='coerce').fillna(0)
                    df_stock[stock_qty_col] = pd.to_numeric(df_stock[stock_qty_col], errors='coerce').fillna(0)
                    
                    # Calculate required stock for target days
                    req_1 = (df_stock["SALES_1_QTY"] / b["s1_days"]) * target_days
                    req_2 = (df_stock["SALES_2_QTY"] / b["s2_days"]) * target_days
                    
                    # Use max requirement to prevent stockouts
                    df_stock["REQ_QTY"] = np.maximum(req_1, req_2).round(0)
                    
                    # Calculate Excess
                    df_stock["EXCESS_QTY"] = df_stock[stock_qty_col] - df_stock["REQ_QTY"]
                    
                    processed_branches[b["name"]] = {
                        "df": df_stock,
                        "gold_code_col": gold_code_col,
                        "item_name_col": stock_name_col,
                        "stock_qty_col": stock_qty_col,
                        "is_main": b["is_main"],
                        "s1_days": b["s1_days"],
                        "s2_days": b["s2_days"]
                    }
                
                # Identify Needs and Excess
                main_branch_name = [name for name, data in processed_branches.items() if data["is_main"]][0]
                main_data = processed_branches[main_branch_name]
                df_main = main_data["df"]
                gold_col_main = main_data["gold_code_col"]
                stock_qty_col_main = main_data["stock_qty_col"]
                s1_days_main = main_data["s1_days"]
                s2_days_main = main_data["s2_days"]
                
                # Items needed by Main Branch (Excess Qty < 0)
                needed_items = df_main[df_main["EXCESS_QTY"] < 0].copy()
                needed_items["NEED_QTY"] = needed_items["EXCESS_QTY"].abs()
                
                # Rename main branch columns
                rename_dict = {
                    "SALES_1_QTY": f"{main_branch_name} {int(s1_days_main)} Days Sales",
                    "SALES_2_QTY": f"{main_branch_name} {int(s2_days_main)} Days Sales",
                    stock_qty_col_main: f"{main_branch_name} Closing Qty",
                    "REQ_QTY": "Required Quantity"
                }
                df_main.rename(columns=rename_dict, inplace=True)
                
                # Merge other branches into df_main
                for name, data in processed_branches.items():
                    if not data["is_main"]:
                        df_other = data["df"]
                        gold_col_other = data["gold_code_col"]
                        stock_qty_col_other = data["stock_qty_col"]
                        s1_days = data["s1_days"]
                        s2_days = data["s2_days"]
                        
                        cols_to_merge = [gold_col_other, "SALES_1_QTY", "SALES_2_QTY", stock_qty_col_other]
                        df_other_sub = df_other[cols_to_merge].copy()
                        df_other_sub = df_other_sub.groupby(gold_col_other).sum().reset_index()
                        
                        df_other_sub.rename(columns={
                            "SALES_1_QTY": f"{name} {int(s1_days)} Days Sales",
                            "SALES_2_QTY": f"{name} {int(s2_days)} Days Sales",
                            stock_qty_col_other: f"{name} Closing Qty"
                        }, inplace=True)
                        
                        df_main = pd.merge(df_main, df_other_sub, left_on=gold_col_main, right_on=gold_col_other, how="left")
                        if gold_col_main != gold_col_other and gold_col_other in df_main.columns:
                            df_main.drop(columns=[gold_col_other], inplace=True)
                            
                # Fill NaNs for the newly merged columns
                for col in df_main.columns:
                    if "Days Sales" in col or "Closing Qty" in col:
                        df_main[col] = df_main[col].fillna(0)
                
                # Consolidate Excess from other branches
                excess_records = []
                for name, data in processed_branches.items():
                    if not data["is_main"]:
                        df_other = data["df"]
                        gold_col = data["gold_code_col"]
                        stock_qty_col_other = data["stock_qty_col"]
                        s2_days = data["s2_days"]
                        
                        # Items with excess > 0
                        excess_items = df_other[df_other["EXCESS_QTY"] > 0]
                        for _, row in excess_items.iterrows():
                            current_stock = row[stock_qty_col_other]
                            sales_s2 = row["SALES_2_QTY"]
                            
                            if sales_s2 > 0:
                                sales_per_day = sales_s2 / s2_days
                                days_coverage = round(current_stock / sales_per_day, 0)
                            else:
                                days_coverage = "No Sales"
                                
                            excess_records.append({
                                "Branch": name,
                                "Item Name": row.get(data["item_name_col"], ""),
                                "Gold Code": row[gold_col],
                                "Item Code": row.get("ITEM CODE", row.get("BARCODE", "")),
                                f"{s2_days} Days Sales": sales_s2,
                                "Current Stock": current_stock,
                                "Available Excess": row["EXCESS_QTY"],
                                "Inventory Coverage (Days)": days_coverage
                            })
                
                df_excess_pool = pd.DataFrame(excess_records)
                
                # Match Needs with Excess
                recommendations = []
                # Keep track of remaining excess
                if not df_excess_pool.empty:
                    available_excess = df_excess_pool.groupby(["Branch", "Gold Code"])["Available Excess"].sum().reset_index()
                else:
                    available_excess = pd.DataFrame(columns=["Branch", "Gold Code", "Available Excess"])
                
                for _, req_row in needed_items.iterrows():
                    g_code = req_row[main_data["gold_code_col"]]
                    needed_qty = req_row["NEED_QTY"]
                    item_name = req_row[main_data["item_name_col"]]
                    
                    # Find branches with excess for this gold code
                    sources = available_excess[(available_excess["Gold Code"] == g_code) & (available_excess["Available Excess"] > 0)].sort_values(by="Available Excess", ascending=False)
                    
                    for idx, src_row in sources.iterrows():
                        if needed_qty <= 0:
                            break
                        
                        take_qty = min(needed_qty, src_row["Available Excess"])
                        recommendations.append({
                            "Gold Code": g_code,
                            "Item Name": item_name,
                            "Source Branch": src_row["Branch"],
                            "Transfer Qty": take_qty
                        })
                        
                        needed_qty -= take_qty
                        available_excess.at[idx, "Available Excess"] -= take_qty

                df_recommendations = pd.DataFrame(recommendations)
                
                # Prepare summary sheet
                summary_data = []
                if not df_recommendations.empty:
                    summary = df_recommendations.groupby("Source Branch")["Transfer Qty"].sum().reset_index()
                    summary.rename(columns={"Transfer Qty": "Total Transfer Qty"}, inplace=True)
                    summary_data = summary
                else:
                    summary_data = pd.DataFrame(columns=["Source Branch", "Total Transfer Qty"])

                # Prepare Summary Dict for Premium Excel
                df_dict = {
                    "Main Branch Data": df_main,
                    "BT Recommendations": df_recommendations if not df_recommendations.empty else pd.DataFrame({"Message": ["No transfers needed"]}),
                    "Branches Excess Stock": df_excess_pool if not df_excess_pool.empty else pd.DataFrame({"Message": ["No excess stock found"]}),
                    "Summary": summary_data if (isinstance(summary_data, pd.DataFrame) and not summary_data.empty) else pd.DataFrame({"Message": ["No transfers"]})
                }

                # Premium Excel Formatting
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    for sheet_name, df_sheet in df_dict.items():
                        df_sheet.to_excel(writer, index=False, sheet_name=sheet_name)
                        worksheet = writer.sheets[sheet_name]
                        
                        header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
                        header_font = Font(bold=True, color="000000")
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                             top=Side(style='thin'), bottom=Side(style='thin'))

                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                                       min_col=1, max_col=worksheet.max_column):
                            for cell in row:
                                cell.border = thin_border
                        
                        for i, col in enumerate(worksheet.columns, 1):
                            max_length = 0
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                                except: pass
                            worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2

                st.success("✅ Analysis & Recommendation Engine Complete!")
                st.download_button(
                    label="📥 Download Premium Branch Transfer Report",
                    data=output.getvalue(),
                    file_name=f"Branch_Transfer_Report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # --- PREMIUM UI UPGRADE: KPIs & Visualizations ---
                st.markdown("---")
                st.markdown("### 📊 Executive Summary & Analytics")
                
                total_need = needed_items["NEED_QTY"].sum() if not needed_items.empty else 0
                total_excess_available = df_excess_pool["Available Excess"].sum() if not df_excess_pool.empty else 0
                total_fulfilled = df_recommendations["Transfer Qty"].sum() if not df_recommendations.empty else 0
                
                # KPI Metrics
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Items in Deficit", len(needed_items))
                m2.metric("Total Qty Needed", f"{total_need:,.0f}")
                m3.metric("Total Qty Sourced", f"{total_fulfilled:,.0f}")
                m4.metric("Fulfillment Rate", f"{(total_fulfilled/total_need*100):.1f}%" if total_need > 0 else "100%")

                # Visualizations
                if not df_recommendations.empty:
                    c1, c2 = st.columns(2)
                    with c1:
                        fig_pie = px.pie(df_recommendations, values="Transfer Qty", names="Source Branch", 
                                         title="Transfers Sourced by Branch", hole=0.4,
                                         color_discrete_sequence=px.colors.sequential.Tealgrn)
                        fig_pie.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="white")
                        st.plotly_chart(fig_pie, use_container_width=True)
                    with c2:
                        top_items = df_recommendations.groupby("Item Name")["Transfer Qty"].sum().nlargest(10).reset_index()
                        fig_bar = px.bar(top_items, x="Transfer Qty", y="Item Name", orientation='h',
                                         title="Top 10 Items Fulfilled", text="Transfer Qty",
                                         color="Transfer Qty", color_continuous_scale="Blues")
                        fig_bar.update_layout(yaxis={'categoryorder':'total ascending'}, 
                                              plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_color="white")
                        st.plotly_chart(fig_bar, use_container_width=True)

                st.markdown("---")
                # Preview
                t1, t2 = st.tabs(["📋 Recommendations Plan", "🔎 Main Branch Deep Dive"])
                with t1:
                    if not df_recommendations.empty:
                        st.dataframe(df_recommendations, use_container_width=True)
                    else:
                        st.info("No recommendations. Main branch is fully stocked or other branches have no excess.")
                with t2:
                    st.dataframe(df_main.head(50), use_container_width=True)
